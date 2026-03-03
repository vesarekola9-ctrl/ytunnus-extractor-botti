# app.py
# Finnish Business Email Finder
# Build: 2026-03-03_clipboard_C_turbo_parallel_cache

import os
import re
import sys
import time
import csv
import threading
import subprocess
from dataclasses import dataclass
from difflib import SequenceMatcher
from typing import Optional, Tuple, List, Dict
from concurrent.futures import ThreadPoolExecutor, as_completed

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import PyPDF2
from docx import Document
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

import requests
import xml.etree.ElementTree as ET

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

import kl_protest_module as klm

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD  # type: ignore
    HAS_DND = True
except Exception:
    HAS_DND = False


APP_BUILD = "2026-03-03_clipboard_C_turbo_parallel_cache"

# --- regex ---
YT_RE = re.compile(r"\b\d{7}-\d\b|\b\d{8}\b")
EMAIL_RE = re.compile(r"[A-Za-z0-9_.+-]+@[A-Za-z0-9-]+\.[A-Za-z0-9-.]+")
EMAIL_A_RE = re.compile(r"[A-Za-z0-9_.+-]+\s*\(a\)\s*[A-Za-z0-9-]+\.[A-Za-z0-9-.]+", re.I)

STRICT_FORMS_RE = re.compile(
    r"\b(oy|ab|ky|tmi|oyj|osakeyhtiö|kommandiittiyhtiö|toiminimi|as\.|ltd|llc|inc|gmbh)\b",
    re.I,
)

KL_PROTEST_DEFAULT_URL = "https://www.kauppalehti.fi/yritykset/protestilista"
YTJ_COMPANY_URL = "https://tietopalvelu.ytj.fi/yritys/{}"

# --- YTJ SOAP company search (name -> Y-tunnus) ---
# Public documentation shows HTTP GET for wmYritysHaku on api.tietopalvelu.ytj.fi :contentReference[oaicite:1]{index=1}
YTJ_SOAP_HTTPGET = "https://api.tietopalvelu.ytj.fi/yritystiedot.asmx/wmYritysHaku"


# =========================
#   SPEED PROFILES (C)
# =========================
@dataclass
class SpeedProfile:
    name: str
    # KL show-more pacing
    kl_scroll_sleep: float
    kl_post_click_sleep: float
    kl_max_passes: int
    # YTJ email Selenium pacing
    ytj_retry_reads: int
    ytj_retry_sleep: float
    ytj_nayta_passes: int
    ytj_per_company_sleep: float
    # timeouts
    page_load_timeout: int
    ytj_page_load_timeout: int
    # C upgrades
    name_workers: int          # SOAP parallel workers
    email_workers: int         # Selenium parallel workers (each worker has its own driver)
    soap_timeout: float        # requests timeout for SOAP
    soap_max_results: int      # read at most N results per name
    turbo_relaxed_match: bool  # allow softer name matching in Turbo


SPEEDS: Dict[str, SpeedProfile] = {
    "Safe": SpeedProfile(
        name="Safe",
        kl_scroll_sleep=0.35,
        kl_post_click_sleep=0.55,
        kl_max_passes=700,
        ytj_retry_reads=8,
        ytj_retry_sleep=0.20,
        ytj_nayta_passes=4,
        ytj_per_company_sleep=0.08,
        page_load_timeout=25,
        ytj_page_load_timeout=25,
        name_workers=6,
        email_workers=1,
        soap_timeout=12.0,
        soap_max_results=25,
        turbo_relaxed_match=False,
    ),
    "Normal": SpeedProfile(
        name="Normal",
        kl_scroll_sleep=0.25,
        kl_post_click_sleep=0.35,
        kl_max_passes=500,
        ytj_retry_reads=6,
        ytj_retry_sleep=0.15,
        ytj_nayta_passes=3,
        ytj_per_company_sleep=0.05,
        page_load_timeout=18,
        ytj_page_load_timeout=18,
        name_workers=10,
        email_workers=2,
        soap_timeout=10.0,
        soap_max_results=30,
        turbo_relaxed_match=False,
    ),
    "Fast": SpeedProfile(
        name="Fast",
        kl_scroll_sleep=0.18,
        kl_post_click_sleep=0.22,
        kl_max_passes=450,
        ytj_retry_reads=5,
        ytj_retry_sleep=0.12,
        ytj_nayta_passes=2,
        ytj_per_company_sleep=0.02,
        page_load_timeout=14,
        ytj_page_load_timeout=14,
        name_workers=16,
        email_workers=3,
        soap_timeout=8.0,
        soap_max_results=40,
        turbo_relaxed_match=False,
    ),
    "Turbo": SpeedProfile(
        name="Turbo",
        kl_scroll_sleep=0.14,
        kl_post_click_sleep=0.18,
        kl_max_passes=420,
        ytj_retry_reads=4,
        ytj_retry_sleep=0.10,
        ytj_nayta_passes=2,
        ytj_per_company_sleep=0.00,
        page_load_timeout=12,
        ytj_page_load_timeout=12,
        name_workers=26,
        email_workers=4,
        soap_timeout=6.0,
        soap_max_results=60,
        turbo_relaxed_match=True,
    ),
}


# =========================
#   SCROLLABLE UI WRAPPER
# =========================
class ScrollableFrame(ttk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        self.canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.inner = ttk.Frame(self.canvas)
        self._win = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _on_inner_configure(self, _event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self._win, width=event.width)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


BaseTk = TkinterDnD.Tk if HAS_DND else tk.Tk


# =========================
#   PATHS (OUTPUT END ONLY)
# =========================
def exe_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def base_output_dir() -> str:
    base = exe_dir()
    try:
        p = os.path.join(base, "_write_test.tmp")
        with open(p, "w", encoding="utf-8") as f:
            f.write("ok")
        os.remove(p)
        return os.path.join(base, "FinnishBusinessEmailFinder")
    except Exception:
        home = os.path.expanduser("~")
        docs = os.path.join(home, "Documents")
        return os.path.join(docs, "FinnishBusinessEmailFinder")


def create_final_run_dir() -> str:
    root = base_output_dir()
    date_folder = time.strftime("%Y-%m-%d")
    run_folder = "run_" + time.strftime("%H-%M-%S")
    out = os.path.join(root, date_folder, run_folder)
    os.makedirs(out, exist_ok=True)
    return out


def open_folder(path: str):
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')
    except Exception:
        pass


# =========================
#   DATA MODEL
# =========================
@dataclass
class Row:
    name: str = ""
    yt: str = ""
    email: str = ""
    source: str = ""
    notes: str = ""


# =========================
#   UTIL
# =========================
def normalize_yt(yt: str) -> Optional[str]:
    yt = (yt or "").strip().replace(" ", "")
    if re.fullmatch(r"\d{7}-\d", yt):
        return yt
    if re.fullmatch(r"\d{8}", yt):
        return yt[:7] + "-" + yt[7]
    return None


def extract_yts_from_text(text: str) -> List[str]:
    yts = set()
    for m in YT_RE.findall(text or ""):
        n = normalize_yt(m)
        if n:
            yts.add(n)
    return sorted(yts)


def pick_email_from_text(text: str) -> str:
    if not text:
        return ""
    m = EMAIL_RE.search(text)
    if m:
        return m.group(0).strip().replace(" ", "")
    m2 = EMAIL_A_RE.search(text)
    if m2:
        return m2.group(0).replace(" ", "").replace("(a)", "@")
    return ""


def split_lines(text: str) -> List[str]:
    if not text:
        return []
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return [ln.strip() for ln in text.split("\n") if ln.strip()]


def extract_names_from_text(text: str, strict: bool, max_names: int) -> List[str]:
    lines = split_lines(text)
    out: List[str] = []
    seen = set()

    bad_contains = [
        "näytä lisää", "kirjaudu", "tilaa", "tilaajille",
        "€", "y-tunnus", "ytunnus", "sähköposti", "puhelin",
        "www.", "http", "kauppalehti", "pörssi", "indeksit"
    ]

    for ln in lines:
        if len(out) >= max_names:
            break
        if YT_RE.search(ln):
            continue
        low = ln.lower()
        if any(b in low for b in bad_contains):
            continue
        if len(ln) < 3:
            continue
        if sum(ch.isdigit() for ch in ln) >= 3:
            continue
        if not any(ch.isalpha() for ch in ln):
            continue

        name = re.sub(r"\s{2,}", " ", ln).strip()
        if len(name) > 90:
            continue
        if strict and not STRICT_FORMS_RE.search(name):
            continue

        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(name)

    return out


def best_name_match_score(query: str, candidate: str) -> float:
    q = (query or "").lower().strip()
    c = (candidate or "").lower().strip()
    if not q or not c:
        return 0.0
    # strong start bonus
    bonus = 0.0
    if c.startswith(q):
        bonus += 18.0
    # basic similarity
    ratio = SequenceMatcher(None, q, c).ratio() * 100.0
    return ratio + bonus


# =========================
#   OUTPUT (END ONLY)
# =========================
def save_emails_docx(out_dir: str, emails: List[str]):
    path = os.path.join(out_dir, "emails.docx")
    doc = Document()
    for e in emails:
        if e:
            doc.add_paragraph(e)
    doc.save(path)
    return path


def save_results_csv(out_dir: str, rows: List[Row]):
    path = os.path.join(out_dir, "results.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Name", "Y-tunnus", "Email", "Source", "Notes"])
        for r in rows:
            w.writerow([r.name, r.yt, r.email, r.source, r.notes])
    return path


def _autosize_columns(ws, max_rows=800):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, max_rows) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(70, max(12, max_len + 2))


def save_results_xlsx(out_dir: str, rows: List[Row], source_label: str, source_url: str = "", speed_name: str = ""):
    path = os.path.join(out_dir, "results.xlsx")
    wb = Workbook()
    headers = ["Name", "Y-tunnus", "Email", "Source", "Notes"]
    header_font = Font(bold=True)

    ws = wb.active
    ws.title = "Results"
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="left")

    for r in rows:
        ws.append([r.name, r.yt, r.email, r.source, r.notes])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    _autosize_columns(ws)

    ws2 = wb.create_sheet("Missing")
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws2.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="left")
    for r in rows:
        if not r.email:
            ws2.append([r.name, r.yt, r.email, r.source, r.notes])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:E{ws2.max_row}"
    _autosize_columns(ws2)

    ws3 = wb.create_sheet("Summary")
    ws3.append(["Build", APP_BUILD])
    ws3.append(["Speed profile", speed_name])
    ws3.append(["Source", source_label])
    if source_url:
        ws3.append(["Source URL / File", source_url])
    ws3.append(["Total rows", len(rows)])
    ws3.append(["Rows with Y-tunnus", sum(1 for r in rows if r.yt)])
    ws3.append(["Rows with Email", sum(1 for r in rows if r.email)])
    ws3.append(["Unique Y-tunnus", len({r.yt for r in rows if r.yt})])
    ws3.append(["Unique Emails", len({r.email.lower() for r in rows if r.email})])
    _autosize_columns(ws3)

    wb.save(path)
    return path


# =========================
#   PDF -> YTs
# =========================
def extract_ytunnukset_from_pdf(pdf_path: str) -> List[str]:
    yt_set = set()
    with open(pdf_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            text = page.extract_text() or ""
            for m in YT_RE.findall(text):
                n = normalize_yt(m)
                if n:
                    yt_set.add(n)
    return sorted(yt_set)


# =========================
#   SELENIUM COMMON
# =========================
def start_new_driver(speed: SpeedProfile):
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    driver_path = ChromeDriverManager().install()
    drv = webdriver.Chrome(service=Service(driver_path), options=options)
    drv.set_page_load_timeout(speed.page_load_timeout)
    return drv


def start_driver_attach_debug(port: int, speed: SpeedProfile):
    options = webdriver.ChromeOptions()
    options.add_experimental_option("debuggerAddress", f"127.0.0.1:{port}")
    driver_path = ChromeDriverManager().install()
    drv = webdriver.Chrome(service=Service(driver_path), options=options)
    drv.set_page_load_timeout(speed.page_load_timeout)
    return drv


def safe_click(driver, elem) -> bool:
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
        time.sleep(0.02)
        try:
            elem.click()
        except Exception:
            driver.execute_script("arguments[0].click();", elem)
        return True
    except Exception:
        return False


def wait_loaded(driver, timeout=18):
    WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.TAG_NAME, "body")))


def try_accept_cookies(driver):
    texts = ["Hyväksy", "Hyväksy kaikki", "Salli kaikki", "Accept", "Accept all", "I agree", "OK", "Selvä"]
    for _ in range(2):
        for e in driver.find_elements(By.XPATH, "//button|//a|//*[@role='button']"):
            try:
                t = (e.text or "").strip()
                if not t:
                    continue
                if any(x.lower() in t.lower() for x in texts):
                    if e.is_displayed() and e.is_enabled():
                        safe_click(driver, e)
                        time.sleep(0.20)
                        break
            except Exception:
                continue


# =========================
#   YTJ EMAIL (Selenium)
# =========================
def extract_email_from_ytj(driver) -> str:
    try:
        for a in driver.find_elements(By.TAG_NAME, "a"):
            href = (a.get_attribute("href") or "")
            if href.lower().startswith("mailto:"):
                return href.split(":", 1)[1].strip()
    except Exception:
        pass

    try:
        cells = driver.find_elements(By.XPATH, "//tr//*[self::td or self::th][contains(normalize-space(.), 'Sähköposti')]")
        for c in cells:
            try:
                tr = c.find_element(By.XPATH, "ancestor::tr[1]")
                email = pick_email_from_text(tr.text or "")
                if email:
                    return email
            except Exception:
                continue
    except Exception:
        pass

    try:
        return pick_email_from_text(driver.find_element(By.TAG_NAME, "body").text or "")
    except Exception:
        return ""


def click_all_nayta_ytj(driver, speed: SpeedProfile):
    for _ in range(speed.ytj_nayta_passes):
        clicked = False
        for b in driver.find_elements(By.TAG_NAME, "button"):
            try:
                if (b.text or "").strip().lower() == "näytä" and b.is_displayed() and b.is_enabled():
                    safe_click(driver, b)
                    clicked = True
                    time.sleep(0.06)
            except Exception:
                continue
        for a in driver.find_elements(By.TAG_NAME, "a"):
            try:
                if (a.text or "").strip().lower() == "näytä" and a.is_displayed():
                    safe_click(driver, a)
                    clicked = True
                    time.sleep(0.06)
            except Exception:
                continue
        if not clicked:
            break


def fetch_email_by_yt(driver, yt: str, stop_flag: threading.Event, speed: SpeedProfile) -> str:
    if stop_flag.is_set():
        return ""
    try:
        driver.get(YTJ_COMPANY_URL.format(yt))
    except TimeoutException:
        pass

    try:
        wait_loaded(driver, timeout=speed.ytj_page_load_timeout)
    except Exception:
        pass
    try_accept_cookies(driver)

    for _ in range(2):
        if stop_flag.is_set():
            return ""
        email = extract_email_from_ytj(driver)
        if email:
            return email
        time.sleep(speed.ytj_retry_sleep)

    click_all_nayta_ytj(driver, speed)
    for _ in range(speed.ytj_retry_reads):
        if stop_flag.is_set():
            return ""
        email = extract_email_from_ytj(driver)
        if email:
            return email
        time.sleep(speed.ytj_retry_sleep)
    return ""


# =========================
#   YTJ SOAP: NAME -> YT (FAST, PARALLEL)
# =========================
def ytj_soap_search_name(name: str, speed: SpeedProfile) -> List[Tuple[str, str]]:
    """
    Returns list of (ytunnus, yritysnimi) from SOAP HTTP GET.
    Public docs list wmYritysHaku over HTTP GET and response schema. :contentReference[oaicite:2]{index=2}
    """
    params = {
        "hakusana": name,
        "yritysmuoto": "",
        "sanahaku": "true",
        "ytunnus": "",
        "voimassaolevat": "true",
        "kieli": "fi",
        "asiakastunnus": "",
        "aikaleima": "",
        "tarkiste": "",
        "tiketti": "",
    }

    r = requests.get(YTJ_SOAP_HTTPGET, params=params, timeout=speed.soap_timeout)
    r.raise_for_status()

    # Parse XML
    # Root has namespace: http://www.ytj.fi/
    xml = r.text
    root = ET.fromstring(xml)

    ns = {"n": "http://www.ytj.fi/"}
    out: List[Tuple[str, str]] = []

    for dto in root.findall(".//n:YritysHakuDTO", ns):
        yt = (dto.findtext("n:YTunnus", default="", namespaces=ns) or "").strip()
        nm = (dto.findtext("n:Yritysnimi", default="", namespaces=ns) or "").strip()
        if yt and nm:
            n_yt = normalize_yt(yt)
            if n_yt:
                out.append((n_yt, nm))

    return out[: max(5, speed.soap_max_results)]


def resolve_name_to_best_yt(name: str, speed: SpeedProfile) -> Tuple[str, str]:
    """
    Returns (yt, matched_name).
    """
    try:
        results = ytj_soap_search_name(name, speed)
    except Exception:
        return "", ""

    if not results:
        return "", ""

    best = ("", "", -1.0)
    for yt, nm in results:
        s = best_name_match_score(name, nm)
        best = (yt, nm, s) if s > best[2] else best

    yt_best, nm_best, score = best

    # Turbo can accept slightly weaker matches
    if speed.turbo_relaxed_match:
        return yt_best, nm_best

    # Normal/Fast/Safe: require a decent score
    if score >= 70.0:
        return yt_best, nm_best
    return "", ""


# =========================
#   PIPELINES (C)
# =========================
def _emails_from_rows(rows: List[Row]) -> List[str]:
    # dedup preserve nice casing
    m = {}
    for r in rows:
        if r.email:
            m[r.email.lower()] = r.email
    return sorted(m.values())


def pipeline_pdf(pdf_path: str, status_cb, progress_cb, stop_flag: threading.Event, speed: SpeedProfile):
    status_cb("PDF: Luetaan ja kerätään Y-tunnukset…")
    yts = extract_ytunnukset_from_pdf(pdf_path)
    if not yts:
        status_cb("PDF: ei löytynyt Y-tunnuksia.")
        return [], []

    status_cb(f"PDF: löytyi {len(yts)} Y-tunnusta. Haetaan emailit YTJ:stä…")

    # Parallel email fetch with multiple drivers (C)
    rows = fetch_emails_parallel(yts, stop_flag, status_cb, progress_cb, speed, source="pdf->ytj")
    return rows, _emails_from_rows(rows)


def pipeline_paste(
    text: str,
    strict: bool,
    max_names: int,
    enable_name_fallback: bool,
    status_cb,
    progress_cb,
    stop_flag: threading.Event,
    speed: SpeedProfile,
):
    status_cb("Paste: poimitaan sähköpostit ja Y-tunnukset…")

    direct_emails = set(e.strip().lower() for e in EMAIL_RE.findall(text or "") if e.strip())
    yts = extract_yts_from_text(text)

    rows: List[Row] = []

    # direct emails
    for em in sorted(direct_emails):
        rows.append(Row(name="", yt="", email=em, source="paste", notes="email found in pasted text"))

    # yts
    for yt in yts:
        rows.append(Row(name="", yt=yt, email="", source="paste->ytj", notes="yt found in pasted text"))

    # name fallback if enabled and NO YTs found
    names: List[str] = []
    if enable_name_fallback and not yts:
        status_cb("Paste: ei Y-tunnuksia – kerätään yritysnimet…")
        names = extract_names_from_text(text, strict=strict, max_names=max_names)
        # dedup
        names = sorted({n.strip(): n.strip() for n in names if n.strip()}.values())

        if not names:
            status_cb("Paste: nimifallback ei löytänyt nimiä.")
        else:
            status_cb(f"Paste: löytyi {len(names)} nimeä. Haetaan Y-tunnukset (SOAP) rinnakkain…")

    # C: resolve names -> YTs in parallel (NO Selenium here)
    name_to_yt: Dict[str, Tuple[str, str]] = {}
    if names:
        name_to_yt = resolve_names_parallel(names, stop_flag, status_cb, progress_cb, speed)

        for nm in names:
            yt, matched = name_to_yt.get(nm, ("", ""))
            if yt:
                rows.append(Row(name=matched or nm, yt=yt, email="", source="paste(name)->ytj", notes=f"name->yt via SOAP | q={nm}"))
            else:
                rows.append(Row(name=nm, yt="", email="", source="paste(name)->ytj", notes="name->yt not found"))

    # dedup rows
    deduped = []
    seen = set()
    for r in rows:
        key = (r.name.lower(), r.yt, r.email.lower())
        if key in seen:
            continue
        seen.add(key)
        deduped.append(r)
    rows = deduped

    if not rows:
        status_cb("Paste: en löytänyt mitään (email / Y-tunnus / nimi).")
        return [], []

    # fetch emails for YTs (parallel)
    yts_to_fetch = sorted({r.yt for r in rows if r.yt and not r.email})
    if not yts_to_fetch:
        status_cb("Paste: valmista (ei YTJ email-hakuja).")
        return rows, _emails_from_rows(rows)

    status_cb(f"YTJ: haetaan emailit ({len(yts_to_fetch)} Y-tunnusta) rinnakkain…")
    fetched_rows = fetch_emails_parallel(yts_to_fetch, stop_flag, status_cb, progress_cb, speed, source="paste->ytj")

    yt_to_email = {r.yt: r.email for r in fetched_rows if r.yt}
    for r in rows:
        if r.yt and not r.email:
            r.email = yt_to_email.get(r.yt, "") or r.email

    return rows, _emails_from_rows(rows)


def pipeline_protest_attach(
    url: str,
    port: int,
    test_limit: int,
    status_cb,
    progress_cb,
    stop_flag: threading.Event,
    speed: SpeedProfile,
):
    status_cb("KL: Yhdistetään Chromeen (debug attach)…")
    driver = start_driver_attach_debug(port, speed)

    yts: List[str] = []
    try:
        status_cb("KL: Avataan protestilista…")
        klm.ensure_on_page(driver, url, status_cb=status_cb)

        status_cb("KL: Ladataan kaikki (Näytä lisää)…")
        klm.click_show_more_until_end(
            driver,
            stop_flag=stop_flag,
            status_cb=status_cb,
            max_passes=speed.kl_max_passes,
            scroll_sleep=speed.kl_scroll_sleep,
            post_click_sleep=speed.kl_post_click_sleep,
        )

        status_cb("KL: Kerätään Y-tunnukset (JS/regex)…")
        yts = klm.extract_ytunnukset_via_js(driver)
        if not yts:
            status_cb("KL: Ei löytynyt Y-tunnuksia. Oletko kirjautunut ja protestilista auki?")
            return [], []
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    if test_limit and test_limit > 0:
        yts = yts[:test_limit]
        status_cb(f"TEST RUN: käsitellään vain {len(yts)} ensimmäistä Y-tunnusta…")
    else:
        status_cb(f"KL: Löytyi {len(yts)} Y-tunnusta. Haetaan YTJ emailit…")

    rows = fetch_emails_parallel(yts, stop_flag, status_cb, progress_cb, speed, source="protest->ytj")
    return rows, _emails_from_rows(rows)


# =========================
#   C: PARALLEL RESOLVE + PARALLEL EMAIL FETCH
# =========================
def resolve_names_parallel(
    names: List[str],
    stop_flag: threading.Event,
    status_cb,
    progress_cb,
    speed: SpeedProfile
) -> Dict[str, Tuple[str, str]]:
    """
    Parallel: name -> (yt, matched_name) using SOAP.
    Cache inside this run.
    """
    out: Dict[str, Tuple[str, str]] = {}
    cache: Dict[str, Tuple[str, str]] = {}

    total = max(1, len(names))
    progress_cb(0, total)

    def worker(nm: str):
        key = nm.strip().lower()
        if key in cache:
            return nm, cache[key]
        yt, matched = resolve_name_to_best_yt(nm, speed)
        cache[key] = (yt, matched)
        return nm, (yt, matched)

    done = 0
    with ThreadPoolExecutor(max_workers=max(1, speed.name_workers)) as ex:
        futures = [ex.submit(worker, nm) for nm in names]
        for fut in as_completed(futures):
            if stop_flag.is_set():
                break
            try:
                nm, res = fut.result()
                out[nm] = res
            except Exception:
                pass
            done += 1
            if done % 5 == 0 or done == len(names):
                status_cb(f"YTJ SOAP: nimihaut {done}/{len(names)}")
            progress_cb(done, total)

    return out


def fetch_emails_parallel(
    yts: List[str],
    stop_flag: threading.Event,
    status_cb,
    progress_cb,
    speed: SpeedProfile,
    source: str
) -> List[Row]:
    """
    Parallel: yt -> email using Selenium, each worker has its own driver.
    Has run-level cache to avoid repeats.
    """
    yts = sorted({yt for yt in yts if yt})
    total = max(1, len(yts))
    progress_cb(0, total)

    cache_email: Dict[str, str] = {}
    rows: List[Row] = []
    lock = threading.Lock()

    # split work among workers
    workers = max(1, speed.email_workers)
    chunks = [[] for _ in range(workers)]
    for i, yt in enumerate(yts):
        chunks[i % workers].append(yt)

    def email_worker(worker_id: int, yt_list: List[str]) -> List[Row]:
        local_rows: List[Row] = []
        if not yt_list:
            return local_rows

        drv = start_new_driver(speed)
        try:
            for yt in yt_list:
                if stop_flag.is_set():
                    break

                with lock:
                    if yt in cache_email:
                        em = cache_email[yt]
                        local_rows.append(Row(name="", yt=yt, email=em, source=source, notes="cache"))
                        continue

                em = fetch_email_by_yt(drv, yt, stop_flag, speed)

                with lock:
                    cache_email[yt] = em or ""
                local_rows.append(Row(name="", yt=yt, email=em, source=source, notes=""))

                if speed.ytj_per_company_sleep > 0:
                    time.sleep(speed.ytj_per_company_sleep)
        finally:
            try:
                drv.quit()
            except Exception:
                pass

        return local_rows

    done = 0
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = []
        for w, chunk in enumerate(chunks):
            futs.append(ex.submit(email_worker, w, chunk))

        for fut in as_completed(futs):
            part = []
            try:
                part = fut.result()
            except Exception:
                part = []
            rows.extend(part)
            # update progress approximately
            done = min(len(cache_email), len(yts))
            status_cb(f"YTJ email: {done}/{len(yts)}")
            progress_cb(done, total)

    # ensure order by yt
    rows.sort(key=lambda r: r.yt)
    progress_cb(len(yts), total)
    return rows


# =========================
#   CHROME DEBUG STARTER
# =========================
def find_chrome_exe() -> Optional[str]:
    candidates = [
        os.path.join(os.environ.get("ProgramFiles", r"C:\Program Files"), r"Google\Chrome\Application\chrome.exe"),
        os.path.join(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)"), r"Google\Chrome\Application\chrome.exe"),
        os.path.join(os.environ.get("LocalAppData", ""), r"Google\Chrome\Application\chrome.exe"),
    ]
    for p in candidates:
        if p and os.path.exists(p):
            return p
    return "chrome.exe"


def default_debug_user_data_dir() -> str:
    return os.path.join(os.path.expanduser("~"), "Documents", "ChromeDebugProfile")


def start_chrome_debug(port: int, user_data_dir: str) -> Tuple[bool, str]:
    chrome = find_chrome_exe()
    if not chrome:
        return False, "Chrome.exe ei löytynyt."
    try:
        os.makedirs(user_data_dir, exist_ok=True)
    except Exception:
        return False, f"Ei voitu luoda user-data-dir: {user_data_dir}"

    args = [chrome, f"--remote-debugging-port={port}", f'--user-data-dir={user_data_dir}']
    try:
        subprocess.Popen(args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, shell=False)
        return True, f"Chrome käynnistetty debug-tilassa porttiin {port}."
    except Exception as e:
        return False, f"Chrome käynnistys epäonnistui: {e}"


# =========================
#   APP UI
# =========================
class App(BaseTk):
    def __init__(self):
        super().__init__()
        self.stop_flag = threading.Event()
        self.last_output_dir: Optional[str] = None

        # theme
        self.BG = "#ffffff"
        self.CARD = "#f6f8ff"
        self.BORDER = "#d7def7"
        self.TEXT = "#0b1b3a"
        self.MUTED = "#465a82"
        self.BLUE = "#1d4ed8"
        self.BLUE_H = "#1e40af"
        self.DANGER = "#b91c1c"
        self.DANGER_H = "#991b1b"
        self.GREY_BTN = "#64748b"
        self.GREY_BTN_H = "#475569"

        self.title("Finnish Business Email Finder")
        self.geometry("1060x1020")
        self.configure(bg=self.BG)

        self._build_ui()

    def _card(self, parent):
        return tk.Frame(parent, bg=self.CARD, highlightthickness=1, highlightbackground=self.BORDER)

    def _btn(self, parent, text, cmd, kind="blue"):
        if kind == "danger":
            bg, hover = self.DANGER, self.DANGER_H
        elif kind == "grey":
            bg, hover = self.GREY_BTN, self.GREY_BTN_H
        else:
            bg, hover = self.BLUE, self.BLUE_H

        b = tk.Button(
            parent, text=text, command=cmd,
            bg=bg, fg="#ffffff",
            activebackground=hover, activeforeground="#ffffff",
            relief="flat", padx=14, pady=10,
            font=("Segoe UI", 10, "bold")
        )
        b.bind("<Enter>", lambda e: b.configure(bg=hover))
        b.bind("<Leave>", lambda e: b.configure(bg=bg))
        return b

    def _ui_log(self, msg):
        ts = time.strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        try:
            self.listbox.insert(tk.END, line)
            self.listbox.yview_moveto(1.0)
        except Exception:
            pass

    def _set_status(self, s):
        self.status.config(text=s)
        self.update_idletasks()
        self._ui_log(s)

    def _set_progress(self, v, mx):
        self.progress["maximum"] = mx
        self.progress["value"] = v
        self.update_idletasks()

    def request_stop(self):
        self.stop_flag.set()
        self._set_status("Pysäytetään…")

    def _clear_stop(self):
        self.stop_flag.clear()

    def clear_paste_text(self):
        self.paste_text.delete("1.0", tk.END)
        self._ui_log("Kenttä tyhjennetty.")

    def open_last_output(self):
        if self.last_output_dir and os.path.isdir(self.last_output_dir):
            open_folder(self.last_output_dir)
        else:
            messagebox.showinfo("Ei kansiota", "Ei vielä tuloskansiota (ajo ei valmis / ei tuloksia).")

    def _current_speed(self) -> SpeedProfile:
        name = (self.speed_var.get() or "Normal").strip()
        return SPEEDS.get(name, SPEEDS["Normal"])

    def _build_ui(self):
        outer = ScrollableFrame(self)
        outer.pack(fill="both", expand=True)
        root = outer.inner

        header = tk.Frame(root, bg=self.BG)
        header.pack(fill="x", padx=16, pady=(16, 10))

        tk.Label(header, text="Finnish Business Email Finder", bg=self.BG, fg=self.TEXT,
                 font=("Segoe UI", 20, "bold")).pack(anchor="w")
        tk.Label(header, text=f"Build: {APP_BUILD}", bg=self.BG, fg=self.MUTED, font=("Segoe UI", 9)).pack(anchor="w")

        actions = self._card(root)
        actions.pack(fill="x", padx=16, pady=(0, 12))
        row = tk.Frame(actions, bg=self.CARD)
        row.pack(fill="x", padx=12, pady=12)

        self._btn(row, "PLAY: Protestilista → YTJ", self.start_protest_mode).pack(side="left", padx=6)
        self._btn(row, "Paste/Clipboard → YTJ", self.start_paste_mode).pack(side="left", padx=6)
        self._btn(row, "PDF → YTJ", self.start_pdf_mode).pack(side="left", padx=6)

        self._btn(row, "Avaa tuloskansio", self.open_last_output, kind="grey").pack(side="right", padx=6)
        self._btn(row, "Pysäytä", self.request_stop, kind="danger").pack(side="right", padx=6)

        status_card = self._card(root)
        status_card.pack(fill="x", padx=16, pady=(0, 12))
        self.status = tk.Label(status_card, text="Valmiina.", bg=self.CARD, fg=self.TEXT, font=("Segoe UI", 11))
        self.status.pack(anchor="w", padx=12, pady=(12, 6))
        self.progress = ttk.Progressbar(status_card, orient="horizontal", mode="determinate", length=920)
        self.progress.pack(fill="x", padx=12, pady=(0, 12))

        # PROTEST PLAY CARD
        play_card = self._card(root)
        play_card.pack(fill="x", padx=16, pady=(0, 12))

        top = tk.Frame(play_card, bg=self.CARD)
        top.pack(fill="x", padx=12, pady=(12, 6))

        tk.Label(top, text="Protestilista URL:", bg=self.CARD, fg=self.TEXT, font=("Segoe UI", 10, "bold")).pack(side="left")
        self.protest_url_var = tk.StringVar(value=KL_PROTEST_DEFAULT_URL)
        tk.Entry(top, textvariable=self.protest_url_var, width=62, bg="#ffffff", fg=self.TEXT,
                 highlightthickness=1, highlightbackground=self.BORDER).pack(side="left", padx=10)

        tk.Label(top, text="Portti:", bg=self.CARD, fg=self.TEXT, font=("Segoe UI", 10)).pack(side="left", padx=(8, 6))
        self.debug_port_var = tk.IntVar(value=9222)
        tk.Spinbox(top, from_=1024, to=65535, textvariable=self.debug_port_var, width=7,
                   bg="#ffffff", fg=self.TEXT, insertbackground=self.TEXT,
                   highlightthickness=1, highlightbackground=self.BORDER).pack(side="left")

        top2 = tk.Frame(play_card, bg=self.CARD)
        top2.pack(fill="x", padx=12, pady=(0, 12))

        tk.Label(top2, text="Nopeus:", bg=self.CARD, fg=self.TEXT, font=("Segoe UI", 10)).pack(side="left")
        self.speed_var = tk.StringVar(value="Normal")
        ttk.Combobox(top2, textvariable=self.speed_var, values=list(SPEEDS.keys()), width=10, state="readonly").pack(side="left", padx=8)

        tk.Label(top2, text="Test run:", bg=self.CARD, fg=self.TEXT, font=("Segoe UI", 10)).pack(side="left", padx=(16, 6))
        self.test_var = tk.StringVar(value="Full")
        ttk.Combobox(top2, textvariable=self.test_var, values=["Full", "5", "10", "25"], width=6, state="readonly").pack(side="left")

        self._btn(top2, "Käynnistä Chrome debug", self.launch_chrome_debug, kind="grey").pack(side="right", padx=6)

        tk.Label(
            play_card,
            text="Ohje: 1) Käynnistä Chrome debug-tilaan  2) Kirjaudu Kauppalehteen  3) Avaa protestilista  4) Paina PLAY.",
            bg=self.CARD, fg=self.MUTED, font=("Segoe UI", 9)
        ).pack(anchor="w", padx=12, pady=(0, 12))

        # PASTE CARD
        paste_card = self._card(root)
        paste_card.pack(fill="x", padx=16, pady=(0, 12))

        top = tk.Frame(paste_card, bg=self.CARD)
        top.pack(fill="x", padx=12, pady=(12, 6))
        tk.Label(top, text="Ohje:", bg=self.CARD, fg=self.TEXT, font=("Segoe UI", 10, "bold")).pack(side="left")
        tk.Label(top, text="Liitä: Y-tunnuksia / sähköposteja / yritysnimiä (yksi per rivi) → Start.",
                 bg=self.CARD, fg=self.MUTED, font=("Segoe UI", 9)).pack(side="left", padx=10)

        top2 = tk.Frame(paste_card, bg=self.CARD)
        top2.pack(fill="x", padx=12, pady=(0, 6))

        self.strict_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            top2, text="Tiukka nimifallback (vain Oy/Ab/Ky/Tmi/...)",
            variable=self.strict_var,
            bg=self.CARD, fg=self.TEXT,
            selectcolor="#ffffff",
            activebackground=self.CARD, activeforeground=self.TEXT
        ).pack(side="left")

        self.enable_name_fallback_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            top2, text="Käytä nimihakua jos EI löydy Y-tunnuksia (suositus)",
            variable=self.enable_name_fallback_var,
            bg=self.CARD, fg=self.TEXT,
            selectcolor="#ffffff",
            activebackground=self.CARD, activeforeground=self.TEXT
        ).pack(side="left", padx=(16, 0))

        tk.Label(top2, text="Max nimeä:", bg=self.CARD, fg=self.TEXT, font=("Segoe UI", 10)).pack(side="left", padx=(16, 6))
        self.max_names_var = tk.IntVar(value=800)
        tk.Spinbox(top2, from_=10, to=20000, textvariable=self.max_names_var, width=7,
                   bg="#ffffff", fg=self.TEXT, insertbackground=self.TEXT,
                   highlightthickness=1, highlightbackground=self.BORDER).pack(side="left")

        self._btn(top2, "Tyhjennä", self.clear_paste_text, kind="grey").pack(side="right", padx=6)

        tk.Label(paste_card, text="Liitä teksti tähän (Ctrl+V):", bg=self.CARD, fg=self.MUTED,
                 font=("Segoe UI", 10)).pack(anchor="w", padx=12, pady=(6, 6))

        self.paste_text = tk.Text(paste_card, height=10, wrap="word",
                                  bg="#ffffff", fg=self.TEXT, insertbackground=self.TEXT,
                                  highlightthickness=1, highlightbackground=self.BORDER)
        self.paste_text.pack(fill="x", padx=12, pady=(0, 12))

        # PDF CARD
        pdf_card = self._card(root)
        pdf_card.pack(fill="x", padx=16, pady=(0, 12))
        self.drop_var = tk.StringVar(value="PDF: Pudota tähän (tai paina PDF → YTJ ja valitse tiedosto)")
        drop = tk.Label(pdf_card, textvariable=self.drop_var,
                        bg="#ffffff", fg=self.MUTED, font=("Segoe UI", 10),
                        padx=12, pady=10,
                        highlightthickness=1, highlightbackground=self.BORDER)
        drop.pack(fill="x", padx=12, pady=12)
        if HAS_DND:
            drop.drop_target_register(DND_FILES)
            drop.dnd_bind("<<Drop>>", self._on_drop_pdf)

        # LOG
        log_card = self._card(root)
        log_card.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        tk.Label(log_card, text="Logi:", bg=self.CARD, fg=self.MUTED, font=("Segoe UI", 10)).pack(anchor="w", padx=12, pady=(12, 6))

        body = tk.Frame(log_card, bg=self.CARD)
        body.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        self.listbox = tk.Listbox(body, height=14, bg="#ffffff", fg=self.TEXT,
                                  highlightthickness=1, highlightbackground=self.BORDER,
                                  selectbackground="#dbeafe")
        self.listbox.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(body, orient="vertical", command=self.listbox.yview)
        sb.pack(side="right", fill="y")
        self.listbox.configure(yscrollcommand=sb.set)

        self._ui_log("Ready.")

    # ===== Chrome debug launcher =====
    def launch_chrome_debug(self):
        port = int(self.debug_port_var.get() or 9222)
        user_data = default_debug_user_data_dir()
        ok, msg = start_chrome_debug(port, user_data)
        if ok:
            messagebox.showinfo("Chrome Debug", msg + "\n\nKirjaudu Kauppalehteen ja avaa protestilista, sitten PLAY.")
            self._set_status(msg)
        else:
            messagebox.showerror("Chrome Debug", msg)

    # ===== Protest PLAY =====
    def start_protest_mode(self):
        self._clear_stop()
        url = (self.protest_url_var.get() or "").strip()
        if not url:
            messagebox.showwarning("URL puuttuu", "Anna protestilistan URL.")
            return
        port = int(self.debug_port_var.get() or 9222)

        test_raw = (self.test_var.get() or "Full").strip()
        test_limit = 0 if test_raw == "Full" else int(test_raw)

        speed = self._current_speed()
        threading.Thread(target=self._run_protest, args=(url, port, test_limit, speed), daemon=True).start()

    def _run_protest(self, url: str, port: int, test_limit: int, speed: SpeedProfile):
        try:
            self._set_status(f"PLAY: Aloitetaan protestilista → YTJ ({speed.name}) …")
            rows, emails = pipeline_protest_attach(url, port, test_limit, self._set_status, self._set_progress, self.stop_flag, speed)

            if self.stop_flag.is_set():
                self._set_status("Pysäytetty.")
                return
            if not rows:
                self._set_status("Ei löytynyt tuloksia.")
                return

            out_dir = create_final_run_dir()
            self.last_output_dir = out_dir
            save_results_xlsx(out_dir, rows, source_label="Protestilista", source_url=url, speed_name=speed.name)
            save_results_csv(out_dir, rows)
            save_emails_docx(out_dir, emails)

            messagebox.showinfo("Valmis", f"Valmis!\n\nKansio:\n{out_dir}\n\nRivejä: {len(rows)}\nSähköposteja: {len(emails)}")
            self._set_status("Valmis!")
        except Exception as e:
            self._ui_log(f"VIRHE: {e}")
            messagebox.showerror("Virhe", f"Tuli virhe:\n\n{e}")

    # ===== Paste =====
    def start_paste_mode(self):
        self._clear_stop()
        text = self.paste_text.get("1.0", tk.END).strip()
        if not text:
            messagebox.showwarning("Tyhjä", "Liitä ensin teksti (Ctrl+V) kenttään.")
            return
        speed = self._current_speed()
        threading.Thread(target=self._run_paste, args=(text, speed), daemon=True).start()

    def _run_paste(self, text: str, speed: SpeedProfile):
        try:
            self._set_status(f"Paste: Aloitetaan ajo ({speed.name})…")
            strict = bool(self.strict_var.get())
            max_names = int(self.max_names_var.get() or 800)
            enable_name_fallback = bool(self.enable_name_fallback_var.get())

            rows, emails = pipeline_paste(
                text, strict, max_names, enable_name_fallback,
                self._set_status, self._set_progress, self.stop_flag, speed
            )

            if self.stop_flag.is_set():
                self._set_status("Pysäytetty.")
                return
            if not rows:
                self._set_status("Ei löytynyt tuloksia.")
                return

            out_dir = create_final_run_dir()
            self.last_output_dir = out_dir
            save_results_xlsx(out_dir, rows, source_label="Paste/Clipboard", speed_name=speed.name)
            save_results_csv(out_dir, rows)
            save_emails_docx(out_dir, emails)

            messagebox.showinfo("Valmis", f"Valmis!\n\nKansio:\n{out_dir}\n\nRivejä: {len(rows)}\nSähköposteja: {len(emails)}")
            self._set_status("Valmis!")
        except Exception as e:
            self._ui_log(f"VIRHE: {e}")
            messagebox.showerror("Virhe", f"Tuli virhe:\n\n{e}")

    # ===== PDF =====
    def _on_drop_pdf(self, event):
        path = (event.data or "").strip()
        if path.startswith("{") and path.endswith("}"):
            path = path[1:-1]
        if path.lower().endswith(".pdf") and os.path.exists(path):
            self.drop_var.set(f"PDF valittu: {path}")
            self._clear_stop()
            speed = self._current_speed()
            threading.Thread(target=self._run_pdf, args=(path, speed), daemon=True).start()
        else:
            messagebox.showwarning("Ei PDF", "Pudotettu tiedosto ei ollut .pdf")

    def start_pdf_mode(self):
        self._clear_stop()
        path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if path:
            self.drop_var.set(f"PDF valittu: {path}")
            speed = self._current_speed()
            threading.Thread(target=self._run_pdf, args=(path, speed), daemon=True).start()

    def _run_pdf(self, pdf_path: str, speed: SpeedProfile):
        try:
            self._set_status(f"PDF: Aloitetaan ajo ({speed.name})…")
            rows, emails = pipeline_pdf(pdf_path, self._set_status, self._set_progress, self.stop_flag, speed)

            if self.stop_flag.is_set():
                self._set_status("Pysäytetty.")
                return
            if not rows:
                self._set_status("Ei löytynyt tuloksia.")
                return

            out_dir = create_final_run_dir()
            self.last_output_dir = out_dir
            save_results_xlsx(out_dir, rows, source_label="PDF", source_url=pdf_path, speed_name=speed.name)
            save_results_csv(out_dir, rows)
            save_emails_docx(out_dir, emails)

            messagebox.showinfo("Valmis", f"Valmis!\n\nKansio:\n{out_dir}\n\nRivejä: {len(rows)}\nSähköposteja: {len(emails)}")
            self._set_status("Valmis!")
        except Exception as e:
            self._ui_log(f"VIRHE: {e}")
            messagebox.showerror("Virhe", f"Tuli virhe:\n\n{e}")


if __name__ == "__main__":
    App().mainloop()
