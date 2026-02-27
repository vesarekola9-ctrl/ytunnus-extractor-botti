# app.py
# Finnish Business Email Finder
# - PDF -> YTJ (emails)
# - Clipboard -> (names / yt / emails) -> YTJ (emails)
#
# OUTPUT:
#   Creates output folder ONLY WHEN RESULTS ARE SAVED (no empty "log folders")
#   Folder: FinnishBusinessEmailFinder/YYYY-MM-DD/run_HH-MM-SS/
#   Files (ONLY):
#     - emails.docx
#     - results.xlsx  (Results + Found Only)
#     - results.csv
#
# UI:
#   - Scroll whole window
#   - Clipboard "Tyhjennä" + "Pysäytä"
#   - "Avaa tuloskansio" + "Kopioi sähköpostit"
#
# Selenium:
#   - webdriver.ChromeOptions() (no Options import)
#   - YTJ name search selects correct company search field (not y-tunnus-only)
#
# NOTE: No log.txt is written.

import os
import re
import sys
import time
import csv
import threading
import subprocess
from difflib import SequenceMatcher

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import PyPDF2
from docx import Document
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD  # type: ignore
    HAS_DND = True
except Exception:
    HAS_DND = False


APP_BUILD = "2026-02-28_no_logfile_csv_foundonly_lazy_output"

# --- tuning ---
YTJ_PAGE_LOAD_TIMEOUT = 18
YTJ_RETRY_READS = 5
YTJ_RETRY_SLEEP = 0.12
YTJ_NAYTA_PASSES = 2
YTJ_PER_COMPANY_SLEEP = 0.02

NAME_SEARCH_TIMEOUT = 10.0
NAME_SEARCH_SLEEP = 0.18

# --- regex ---
YT_RE = re.compile(r"\b\d{7}-\d\b|\b\d{8}\b")
EMAIL_RE = re.compile(r"[A-Za-z0-9_.+-]+@[A-Za-z0-9-]+\.[A-Za-z0-9-.]+")
EMAIL_A_RE = re.compile(r"[A-Za-z0-9_.+-]+\s*\(a\)\s*[A-Za-z0-9-]+\.[A-Za-z0-9-.]+", re.I)

YTJ_COMPANY_URL = "https://tietopalvelu.ytj.fi/yritys/{}"
YTJ_SEARCH_URLS = ["https://tietopalvelu.ytj.fi/haku", "https://tietopalvelu.ytj.fi/"]

STRICT_FORMS_RE = re.compile(
    r"\b(oy|ab|ky|tmi|oyj|osakeyhtiö|kommandiittiyhtiö|toiminimi|as\.|ltd|llc|inc|gmbh)\b",
    re.I,
)


# =========================
#   SCROLLABLE UI WRAPPER
# =========================
class ScrollableFrame(ttk.Frame):
    """Canvas-based scrollable frame. Mouse wheel scrolls the whole UI."""
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)

        self.canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.inner = ttk.Frame(self.canvas)
        self._win = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)  # Windows

    def _on_inner_configure(self, _event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self._win, width=event.width)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


BaseTk = TkinterDnD.Tk if HAS_DND else tk.Tk


# =========================
#   PATHS / RUN FOLDERS
# =========================
def exe_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def base_output_dir() -> str:
    base = exe_dir()
    try:
        p = os.path.join(base, "_write_test.tmp")
        with open(p, "w", encoding="utf-8") as f:
            f.write("ok")
        os.remove(p)
        return os.path.join(base, "FinnishBusinessEmailFinder")
    except Exception:
        home = os.path.expanduser("~")
        docs = os.path.join(home, "Documents")
        return os.path.join(docs, "FinnishBusinessEmailFinder")


def create_run_dir() -> str:
    root = base_output_dir()
    date_folder = time.strftime("%Y-%m-%d")
    run_folder = "run_" + time.strftime("%H-%M-%S")
    out = os.path.join(root, date_folder, run_folder)
    os.makedirs(out, exist_ok=True)
    return out


def open_path_in_os(path: str):
    if not path:
        return
    try:
        if os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", path], check=False)
        else:
            subprocess.run(["xdg-open", path], check=False)
    except Exception:
        pass


# =========================
#   UTIL
# =========================
def normalize_yt(yt: str):
    yt = (yt or "").strip().replace(" ", "")
    if re.fullmatch(r"\d{7}-\d", yt):
        return yt
    if re.fullmatch(r"\d{8}", yt):
        return yt[:7] + "-" + yt[7]
    return None


def pick_email_from_text(text: str) -> str:
    if not text:
        return ""
    m = EMAIL_RE.search(text)
    if m:
        return m.group(0).strip().replace(" ", "")
    m2 = EMAIL_A_RE.search(text)
    if m2:
        return m2.group(0).replace(" ", "").replace("(a)", "@")
    return ""


def extract_yts_from_text(text: str):
    yts = set()
    for m in YT_RE.findall(text or ""):
        n = normalize_yt(m)
        if n:
            yts.add(n)
    return sorted(yts)


def split_lines(text: str):
    if not text:
        return []
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return [ln.strip() for ln in text.split("\n") if ln.strip()]


def extract_names_from_text(text: str, strict: bool, max_names: int):
    lines = split_lines(text)
    out = []
    seen = set()
    bad_contains = [
        "näytä lisää", "kirjaudu", "tilaa", "tilaajille",
        "€", "y-tunnus", "ytunnus", "sähköposti", "puhelin",
        "www.", "http", "kauppalehti", "protestilista",
        "viimeisimmät protestit", "häiriöpäivä", "velkomustuomiot",
        "sijainti", "summa", "lähde"
    ]

    for ln in lines:
        if len(out) >= max_names:
            break
        if YT_RE.search(ln):
            continue
        low = ln.lower()
        if any(b in low for b in bad_contains):
            continue
        if len(ln) < 3:
            continue
        if sum(ch.isdigit() for ch in ln) >= 3:
            continue
        if not any(ch.isalpha() for ch in ln):
            continue
        name = re.sub(r"\s{2,}", " ", ln).strip()
        if len(name) > 90:
            continue
        if strict and not STRICT_FORMS_RE.search(name):
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(name)
    return out


def extract_yt_from_text_anywhere(txt: str) -> str:
    if not txt:
        return ""
    for m in YT_RE.findall(txt):
        n = normalize_yt(m)
        if n:
            return n
    return ""


# =========================
#   OUTPUT (ONLY at the end)
# =========================
def save_emails_docx(out_dir: str, emails: list[str]):
    path = os.path.join(out_dir, "emails.docx")
    doc = Document()
    for e in emails:
        if e:
            doc.add_paragraph(e)
    doc.save(path)
    return path


def _autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 500) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(60, max(12, max_len + 2))


def save_results_xlsx(out_dir: str, rows: list[dict], filename="results.xlsx"):
    path = os.path.join(out_dir, filename)
    wb = Workbook()

    headers = ["Name", "Y-tunnus", "Email", "Source", "Notes"]

    # Sheet 1: All
    ws = wb.active
    ws.title = "Results"
    ws.append(headers)
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="left")

    for r in rows:
        ws.append([r.get("name", ""), r.get("yt", ""), r.get("email", ""), r.get("source", ""), r.get("notes", "")])

    _autosize_columns(ws)

    # Sheet 2: Found Only
    ws2 = wb.create_sheet("Found Only")
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws2.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="left")

    for r in rows:
        if (r.get("email") or "").strip():
            ws2.append([r.get("name", ""), r.get("yt", ""), r.get("email", ""), r.get("source", ""), r.get("notes", "")])

    _autosize_columns(ws2)

    wb.save(path)
    return path


def save_results_csv(out_dir: str, rows: list[dict], filename="results.csv"):
    path = os.path.join(out_dir, filename)
    headers = ["Name", "Y-tunnus", "Email", "Source", "Notes"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(headers)
        for r in rows:
            w.writerow([r.get("name", ""), r.get("yt", ""), r.get("email", ""), r.get("source", ""), r.get("notes", "")])
    return path


# =========================
#   PDF -> YTs
# =========================
def extract_ytunnukset_from_pdf(pdf_path: str):
    yt_set = set()
    with open(pdf_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            text = page.extract_text() or ""
            for m in YT_RE.findall(text):
                n = normalize_yt(m)
                if n:
                    yt_set.add(n)
    return sorted(yt_set)


# =========================
#   SELENIUM
# =========================
def start_new_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")

    driver_path = ChromeDriverManager().install()
    drv = webdriver.Chrome(service=Service(driver_path), options=options)
    drv.set_page_load_timeout(YTJ_PAGE_LOAD_TIMEOUT)
    return drv


def safe_click(driver, elem) -> bool:
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
        time.sleep(0.01)
        try:
            elem.click()
        except Exception:
            driver.execute_script("arguments[0].click();", elem)
        return True
    except Exception:
        return False


def try_accept_cookies(driver):
    texts = ["Hyväksy", "Hyväksy kaikki", "Salli kaikki", "Accept", "Accept all", "I agree", "OK", "Selvä"]
    for _ in range(2):
        for e in driver.find_elements(By.XPATH, "//button|//a|//*[@role='button']"):
            try:
                t = (e.text or "").strip()
                if not t:
                    continue
                if any(x.lower() in t.lower() for x in texts):
                    if e.is_displayed() and e.is_enabled():
                        safe_click(driver, e)
                        time.sleep(0.15)
                        break
            except Exception:
                continue


def wait_ytj_loaded(driver):
    WebDriverWait(driver, YTJ_PAGE_LOAD_TIMEOUT).until(
        EC.presence_of_element_located((By.TAG_NAME, "body"))
    )


def extract_email_from_ytj(driver):
    try:
        for a in driver.find_elements(By.TAG_NAME, "a"):
            href = (a.get_attribute("href") or "")
            if href.lower().startswith("mailto:"):
                return href.split(":", 1)[1].strip()
    except Exception:
        pass

    try:
        cells = driver.find_elements(By.XPATH, "//tr//*[self::td or self::th][contains(normalize-space(.), 'Sähköposti')]")
        for c in cells:
            try:
                tr = c.find_element(By.XPATH, "ancestor::tr[1]")
                email = pick_email_from_text(tr.text or "")
                if email:
                    return email
            except Exception:
                continue
    except Exception:
        pass

    try:
        return pick_email_from_text(driver.find_element(By.TAG_NAME, "body").text or "")
    except Exception:
        return ""


def click_all_nayta_ytj(driver):
    for _ in range(YTJ_NAYTA_PASSES):
        clicked = False
        for b in driver.find_elements(By.TAG_NAME, "button"):
            try:
                if (b.text or "").strip().lower() == "näytä" and b.is_displayed() and b.is_enabled():
                    safe_click(driver, b)
                    clicked = True
                    time.sleep(0.06)
            except Exception:
                continue
        for a in driver.find_elements(By.TAG_NAME, "a"):
            try:
                if (a.text or "").strip().lower() == "näytä" and a.is_displayed():
                    safe_click(driver, a)
                    clicked = True
                    time.sleep(0.06)
            except Exception:
                continue
        if not clicked:
            break


def fetch_email_by_yt(driver, yt: str, stop_flag: threading.Event):
    if stop_flag.is_set():
        return ""
    try:
        driver.get(YTJ_COMPANY_URL.format(yt))
    except TimeoutException:
        pass
    try:
        wait_ytj_loaded(driver)
    except Exception:
        pass
    try_accept_cookies(driver)

    email = ""
    for _ in range(2):
        if stop_flag.is_set():
            return ""
        email = extract_email_from_ytj(driver)
        if email:
            return email
        time.sleep(YTJ_RETRY_SLEEP)

    click_all_nayta_ytj(driver)
    for _ in range(YTJ_RETRY_READS):
        if stop_flag.is_set():
            return ""
        email = extract_email_from_ytj(driver)
        if email:
            return email
        time.sleep(YTJ_RETRY_SLEEP)
    return ""


# =========================
#   YTJ NAME SEARCH (correct field)
# =========================
def _attr(el, name: str) -> str:
    try:
        return (el.get_attribute(name) or "").strip()
    except Exception:
        return ""


def find_ytj_company_search_input(driver):
    candidates = []
    try:
        inputs = driver.find_elements(By.XPATH, "//input")
    except Exception:
        inputs = []

    for inp in inputs:
        try:
            if not inp.is_displayed() or not inp.is_enabled():
                continue

            itype = (_attr(inp, "type") or "").lower()
            if itype in ("hidden", "password", "checkbox", "radio", "submit", "button", "file"):
                continue

            ph = _attr(inp, "placeholder").lower()
            al = _attr(inp, "aria-label").lower()
            nm = _attr(inp, "name").lower()
            iid = _attr(inp, "id").lower()
            text = " ".join([ph, al, nm, iid]).strip()

            # reject y-tunnus-only inputs
            if "y-tunnus" in text and ("yritys" not in text and "toiminimi" not in text and "nimi" not in text):
                continue

            score = 0
            if "yritys" in text:
                score += 50
            if "toiminimi" in text:
                score += 35
            if "nimi" in text:
                score += 25
            if itype == "search":
                score += 15
            if "hae" in text:
                score += 10
            if "y-tunnus" in text:
                score -= 5

            if score <= 0:
                continue

            candidates.append((score, inp))
        except Exception:
            continue

    if not candidates:
        try:
            for inp in driver.find_elements(By.XPATH, "//input[@type='search']"):
                if inp.is_displayed() and inp.is_enabled():
                    return inp
        except Exception:
            pass
        return None

    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def ytj_open_search_home(driver):
    for url in YTJ_SEARCH_URLS:
        try:
            driver.get(url)
            wait_ytj_loaded(driver)
            try_accept_cookies(driver)
            if find_ytj_company_search_input(driver):
                return True
        except Exception:
            continue
    return False


def score_result(name_query: str, card_text: str) -> float:
    txt = (card_text or "").strip()
    ratio = SequenceMatcher(None, (name_query or "").lower(), txt.lower()).ratio()
    score = ratio * 100.0
    if extract_yt_from_text_anywhere(txt):
        score += 20.0
    return score


def ytj_name_to_yt(driver, name: str, stop_flag: threading.Event):
    if not ytj_open_search_home(driver):
        return ""
    if stop_flag.is_set():
        return ""

    inp = find_ytj_company_search_input(driver)
    if not inp:
        return ""

    try:
        try:
            inp.clear()
        except Exception:
            pass
        inp.send_keys(name)
        inp.send_keys(u"\ue007")  # ENTER
    except Exception:
        return ""

    best_link = None
    best_score = -1.0
    t0 = time.time()

    while time.time() - t0 < NAME_SEARCH_TIMEOUT:
        if stop_flag.is_set():
            return ""

        candidate_links = []
        for xp in ("//a[contains(@href,'/yritys/')]", "//a[contains(@href,'yritys')]"):
            try:
                candidate_links.extend(driver.find_elements(By.XPATH, xp))
            except Exception:
                pass

        checked = 0
        for a in candidate_links:
            if checked >= 30:
                break
            try:
                if not a.is_displayed():
                    continue
                href = (a.get_attribute("href") or "")
                if not href or "tietopalvelu.ytj.fi" not in href:
                    continue

                try:
                    card = a.find_element(By.XPATH, "ancestor::*[self::li or self::div or self::article][1]")
                    card_text = (card.text or "")
                except Exception:
                    card_text = (a.text or "")

                s = score_result(name, card_text)
                checked += 1
                if s > best_score:
                    best_score = s
                    best_link = a
            except Exception:
                continue

        if best_link:
            break

        time.sleep(NAME_SEARCH_SLEEP)

    if not best_link:
        return ""

    try:
        safe_click(driver, best_link)
        try:
            wait_ytj_loaded(driver)
        except Exception:
            pass
        try_accept_cookies(driver)
        try:
            body = driver.find_element(By.TAG_NAME, "body").text or ""
        except Exception:
            body = ""
        return extract_yt_from_text_anywhere(body)
    except Exception:
        return ""


# =========================
#   PIPELINES
# =========================
def pipeline_pdf(pdf_path: str, status_cb, progress_cb, stop_flag: threading.Event):
    status_cb("Luetaan PDF ja kerätään Y-tunnukset…")
    yts = extract_ytunnukset_from_pdf(pdf_path)
    if not yts:
        status_cb("PDF: ei löytynyt Y-tunnuksia.")
        return [], []

    status_cb(f"PDF: löytyi {len(yts)} Y-tunnusta. Haetaan emailit YTJ:stä…")
    driver = start_new_driver()
    rows = []
    try:
        progress_cb(0, max(1, len(yts)))
        cache = {}
        for i, yt in enumerate(yts, start=1):
            if stop_flag.is_set():
                status_cb("Pysäytetty.")
                break
            progress_cb(i - 1, len(yts))
            status_cb(f"YTJ: {i}/{len(yts)} {yt}")

            email = cache.get(yt)
            if email is None:
                email = fetch_email_by_yt(driver, yt, stop_flag)
                cache[yt] = email

            rows.append({"name": "", "yt": yt, "email": email, "source": "pdf->ytj", "notes": ""})
            time.sleep(YTJ_PER_COMPANY_SLEEP)

        progress_cb(len(yts), max(1, len(yts)))
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    emails = sorted({r["email"].lower(): r["email"] for r in rows if r.get("email")}.values())
    return rows, emails


def pipeline_clipboard(
    text: str,
    strict: bool,
    max_names: int,
    force_name_flow: bool,
    status_cb,
    progress_cb,
    stop_flag: threading.Event
):
    status_cb("Clipboard: poimitaan sähköpostit, Y-tunnukset ja yritysnimet…")

    direct_emails = sorted({e.strip().lower() for e in EMAIL_RE.findall(text or "") if e.strip()})
    yts = extract_yts_from_text(text)
    names = extract_names_from_text(text, strict=strict, max_names=max_names)

    rows: list[dict] = []

    # Keep direct emails as convenience rows (doesn't affect name flow)
    for em in direct_emails:
        rows.append({"name": "", "yt": "", "email": em, "source": "clipboard", "notes": "email found in pasted text"})

    # Main items
    main_items: list[dict] = []
    if force_name_flow and names:
        for nm in names:
            main_items.append({"name": nm, "yt": "", "email": "", "source": "clipboard->ytj", "notes": "name->yt->email"})
    else:
        for yt in yts:
            main_items.append({"name": "", "yt": yt, "email": "", "source": "clipboard->ytj", "notes": "yt found in pasted text"})
        for nm in names:
            main_items.append({"name": nm, "yt": "", "email": "", "source": "clipboard->ytj", "notes": "name found in pasted text"})

    # Dedup main items
    seen = set()
    dedup = []
    for r in main_items:
        key = (r.get("name", "").lower(), r.get("yt", ""))
        if key in seen:
            continue
        seen.add(key)
        dedup.append(r)
    main_items = dedup

    if not rows and not main_items:
        status_cb("Clipboard: en löytänyt mitään (email / Y-tunnus / nimi).")
        return [], []

    status_cb("Käynnistetään YTJ-haku…")
    driver = start_new_driver()

    yt_cache: dict[str, str] = {}
    name_cache: dict[str, str] = {}

    try:
        progress_cb(0, max(1, len(main_items)))

        for idx, r in enumerate(main_items, start=1):
            if stop_flag.is_set():
                status_cb("Pysäytetty.")
                break

            progress_cb(idx - 1, len(main_items))
            name = (r.get("name") or "").strip()
            yt = (r.get("yt") or "").strip()

            if yt:
                status_cb(f"YTJ email: {idx}/{len(main_items)} {yt}")
                email = yt_cache.get(yt)
                if email is None:
                    email = fetch_email_by_yt(driver, yt, stop_flag)
                    yt_cache[yt] = email
                r["email"] = email or ""
            else:
                if not name:
                    continue

                status_cb(f"YTJ yrityshaku: {idx}/{len(main_items)} {name}")
                yt2 = name_cache.get(name)
                if yt2 is None:
                    yt2 = ytj_name_to_yt(driver, name, stop_flag)
                    name_cache[name] = yt2

                if yt2:
                    r["yt"] = yt2
                    status_cb(f"YTJ email: {idx}/{len(main_items)} {yt2}")
                    email2 = yt_cache.get(yt2)
                    if email2 is None:
                        email2 = fetch_email_by_yt(driver, yt2, stop_flag)
                        yt_cache[yt2] = email2
                    r["email"] = email2 or ""
                else:
                    r["notes"] = (r.get("notes") or "") + " | yt not found"

            time.sleep(YTJ_PER_COMPANY_SLEEP)

        progress_cb(len(main_items), max(1, len(main_items)))
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    rows.extend(main_items)

    emails = sorted({(r.get("email") or "").lower(): (r.get("email") or "") for r in rows if (r.get("email") or "").strip()}.values())
    return rows, emails


# =========================
#   APP UI
# =========================
class App(BaseTk):
    def __init__(self):
        super().__init__()
        self.stop_flag = threading.Event()
        self.last_emails: list[str] = []
        self.last_out_dir: str = ""

        # theme
        self.BG = "#ffffff"
        self.CARD = "#f6f8ff"
        self.BORDER = "#d7def7"
        self.TEXT = "#0b1b3a"
        self.MUTED = "#465a82"
        self.BLUE = "#1d4ed8"
        self.BLUE_H = "#1e40af"
        self.DANGER = "#b91c1c"
        self.DANGER_H = "#991b1b"
        self.GREY_BTN = "#64748b"
        self.GREY_BTN_H = "#475569"

        self.title("Finnish Business Email Finder")
        self.geometry("980x920")
        self.configure(bg=self.BG)

        self._build_ui()

    def _card(self, parent):
        return tk.Frame(parent, bg=self.CARD, highlightthickness=1, highlightbackground=self.BORDER)

    def _btn(self, parent, text, cmd, kind="blue"):
        if kind == "danger":
            bg, hover = self.DANGER, self.DANGER_H
        elif kind == "grey":
            bg, hover = self.GREY_BTN, self.GREY_BTN_H
        else:
            bg, hover = self.BLUE, self.BLUE_H

        b = tk.Button(
            parent, text=text, command=cmd,
            bg=bg, fg="#ffffff",
            activebackground=hover, activeforeground="#ffffff",
            relief="flat", padx=14, pady=10,
            font=("Segoe UI", 10, "bold")
        )
        b.bind("<Enter>", lambda e: b.configure(bg=hover))
        b.bind("<Leave>", lambda e: b.configure(bg=bg))
        return b

    def _ui_log(self, msg):
        ts = time.strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        try:
            self.listbox.insert(tk.END, line)
            self.listbox.yview_moveto(1.0)
        except Exception:
            pass

    def _set_status(self, s):
        self.status.config(text=s)
        self.update_idletasks()
        self._ui_log(s)

    def _set_progress(self, v, mx):
        self.progress["maximum"] = mx
        self.progress["value"] = v
        self.update_idletasks()

    def request_stop(self):
        self.stop_flag.set()
        self._set_status("Pysäytetään…")

    def _clear_stop(self):
        self.stop_flag.clear()

    def clear_clipboard_text(self):
        self.clip_text.delete("1.0", tk.END)
        self._ui_log("Clipboard tyhjennetty.")

    def open_out_folder(self):
        if not self.last_out_dir or not os.path.isdir(self.last_out_dir):
            messagebox.showinfo("Ei kansiota", "Aja ensin haku. Kansio luodaan vasta lopuksi, kun tulokset tallennetaan.")
            return
        open_path_in_os(self.last_out_dir)

    def copy_emails_to_clipboard(self):
        if not self.last_emails:
            messagebox.showinfo("Ei sähköposteja", "Sähköposteja ei ole vielä listalla. Aja haku ensin.")
            return
        txt = "\n".join(self.last_emails)
        try:
            self.clipboard_clear()
            self.clipboard_append(txt)
            self._ui_log(f"Kopioitu {len(self.last_emails)} sähköpostia leikepöydälle.")
        except Exception:
            messagebox.showwarning("Kopiointi epäonnistui", "En saanut kopioitua leikepöydälle.")

    def _finalize_and_save(self, rows: list[dict], emails: list[str]) -> str:
        # create output folder ONLY at the end
        out_dir = create_run_dir()
        save_results_xlsx(out_dir, rows)
        save_results_csv(out_dir, rows)
        save_emails_docx(out_dir, emails)
        return out_dir

    def _build_ui(self):
        outer = ScrollableFrame(self)
        outer.pack(fill="both", expand=True)
        root = outer.inner

        header = tk.Frame(root, bg=self.BG)
        header.pack(fill="x", padx=16, pady=(16, 10))

        tk.Label(header, text="Finnish Business Email Finder", bg=self.BG, fg=self.TEXT,
                 font=("Segoe UI", 20, "bold")).pack(anchor="w")
        tk.Label(header, text=f"Build: {APP_BUILD}", bg=self.BG, fg=self.MUTED, font=("Segoe UI", 9)).pack(anchor="w")

        actions = self._card(root)
        actions.pack(fill="x", padx=16, pady=(0, 12))
        row = tk.Frame(actions, bg=self.CARD)
        row.pack(fill="x", padx=12, pady=12)

        self._btn(row, "PDF → YTJ", self.start_pdf_mode).pack(side="left", padx=6)
        self._btn(row, "Clipboard → Finder", self.start_clipboard_mode).pack(side="left", padx=6)
        self._btn(row, "Pysäytä", self.request_stop, kind="danger").pack(side="right", padx=6)

        util_row = tk.Frame(actions, bg=self.CARD)
        util_row.pack(fill="x", padx=12, pady=(0, 12))
        self._btn(util_row, "Avaa tuloskansio", self.open_out_folder, kind="grey").pack(side="left", padx=6)
        self._btn(util_row, "Kopioi sähköpostit", self.copy_emails_to_clipboard, kind="grey").pack(side="right", padx=6)

        status_card = self._card(root)
        status_card.pack(fill="x", padx=16, pady=(0, 12))

        self.status = tk.Label(status_card, text="Valmiina.", bg=self.CARD, fg=self.TEXT, font=("Segoe UI", 11))
        self.status.pack(anchor="w", padx=12, pady=(12, 6))

        self.progress = ttk.Progressbar(status_card, orient="horizontal", mode="determinate", length=920)
        self.progress.pack(fill="x", padx=12, pady=(0, 12))

        pdf_card = self._card(root)
        pdf_card.pack(fill="x", padx=16, pady=(0, 12))

        self.drop_var = tk.StringVar(value="PDF: Pudota tähän (tai paina PDF → YTJ ja valitse tiedosto)")
        drop = tk.Label(pdf_card, textvariable=self.drop_var,
                        bg="#ffffff", fg=self.MUTED, font=("Segoe UI", 10),
                        padx=12, pady=10,
                        highlightthickness=1, highlightbackground=self.BORDER)
        drop.pack(fill="x", padx=12, pady=12)

        if HAS_DND:
            drop.drop_target_register(DND_FILES)
            drop.dnd_bind("<<Drop>>", self._on_drop_pdf)

        clip_card = self._card(root)
        clip_card.pack(fill="x", padx=16, pady=(0, 12))

        top = tk.Frame(clip_card, bg=self.CARD)
        top.pack(fill="x", padx=12, pady=(12, 6))

        self.strict_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            top, text="Tiukka parsinta (vain Oy/Ab/Ky/Tmi/...)",
            variable=self.strict_var,
            bg=self.CARD, fg=self.TEXT,
            selectcolor="#ffffff",
            activebackground=self.CARD, activeforeground=self.TEXT
        ).pack(side="left")

        self.force_name_flow_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            top, text="Pakota: Nimi → Y-tunnus → Email (suositus KL-listoille)",
            variable=self.force_name_flow_var,
            bg=self.CARD, fg=self.TEXT,
            selectcolor="#ffffff",
            activebackground=self.CARD, activeforeground=self.TEXT
        ).pack(side="left", padx=(16, 0))

        tk.Label(top, text="Max nimeä:", bg=self.CARD, fg=self.TEXT, font=("Segoe UI", 10)).pack(side="left", padx=(16, 6))
        self.max_names_var = tk.IntVar(value=400)
        tk.Spinbox(top, from_=10, to=5000, textvariable=self.max_names_var, width=7,
                   bg="#ffffff", fg=self.TEXT, insertbackground=self.TEXT,
                   highlightthickness=1, highlightbackground=self.BORDER).pack(side="left")

        self._btn(top, "Tyhjennä", self.clear_clipboard_text, kind="grey").pack(side="right", padx=6)

        tk.Label(clip_card, text="Liitä tähän (Ctrl+V):", bg=self.CARD, fg=self.MUTED,
                 font=("Segoe UI", 10)).pack(anchor="w", padx=12, pady=(6, 6))

        self.clip_text = tk.Text(clip_card, height=9, wrap="word",
                                 bg="#ffffff", fg=self.TEXT, insertbackground=self.TEXT,
                                 highlightthickness=1, highlightbackground=self.BORDER)
        self.clip_text.pack(fill="x", padx=12, pady=(0, 12))

        log_card = self._card(root)
        log_card.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        tk.Label(log_card, text="Logi (vain tässä näkymässä):", bg=self.CARD, fg=self.MUTED, font=("Segoe UI", 10)).pack(anchor="w", padx=12, pady=(12, 6))

        body = tk.Frame(log_card, bg=self.CARD)
        body.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        self.listbox = tk.Listbox(body, height=14, bg="#ffffff", fg=self.TEXT,
                                  highlightthickness=1, highlightbackground=self.BORDER,
                                  selectbackground="#dbeafe")
        self.listbox.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(body, orient="vertical", command=self.listbox.yview)
        sb.pack(side="right", fill="y")
        self.listbox.configure(yscrollcommand=sb.set)

        self._ui_log("Ready.")

    def _on_drop_pdf(self, event):
        path = (event.data or "").strip()
        if path.startswith("{") and path.endswith("}"):
            path = path[1:-1]
        if path.lower().endswith(".pdf") and os.path.exists(path):
            self.drop_var.set(f"PDF valittu: {path}")
            self._clear_stop()
            threading.Thread(target=self._run_pdf, args=(path,), daemon=True).start()
        else:
            messagebox.showwarning("Ei PDF", "Pudotettu tiedosto ei ollut .pdf")

    def start_pdf_mode(self):
        self._clear_stop()
        path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if path:
            self.drop_var.set(f"PDF valittu: {path}")
            threading.Thread(target=self._run_pdf, args=(path,), daemon=True).start()

    def _run_pdf(self, pdf_path):
        self.last_emails = []
        self.last_out_dir = ""
        try:
            self._set_status("Aloitetaan PDF ajo…")
            rows, emails = pipeline_pdf(pdf_path, self._set_status, self._set_progress, self.stop_flag)
            if self.stop_flag.is_set():
                self._set_status("Pysäytetty.")
                return
            if not rows:
                self._set_status("Ei löytynyt tuloksia.")
                return

            self._set_status("Tallennetaan tulokset…")
            out_dir = self._finalize_and_save(rows, emails)
            self.last_out_dir = out_dir
            self.last_emails = emails

            messagebox.showinfo(
                "Valmis",
                f"Valmis!\n\nKansio:\n{out_dir}\n\nRivejä: {len(rows)}\nSähköposteja: {len(emails)}"
            )
            self._set_status("Valmis!")
        except Exception as e:
            self._ui_log(f"VIRHE: {e}")
            messagebox.showerror("Virhe", f"Tuli virhe:\n\n{e}")

    def start_clipboard_mode(self):
        self._clear_stop()
        text = self.clip_text.get("1.0", tk.END).strip()
        if not text:
            messagebox.showwarning("Tyhjä", "Liitä ensin teksti (Ctrl+V) kenttään.")
            return
        threading.Thread(target=self._run_clipboard, args=(text,), daemon=True).start()

    def _run_clipboard(self, text: str):
        self.last_emails = []
        self.last_out_dir = ""
        try:
            self._set_status("Aloitetaan Clipboard ajo…")
            strict = bool(self.strict_var.get())
            max_names = int(self.max_names_var.get() or 400)
            force_name_flow = bool(self.force_name_flow_var.get())

            rows, emails = pipeline_clipboard(
                text, strict, max_names, force_name_flow,
                self._set_status, self._set_progress, self.stop_flag
            )

            if self.stop_flag.is_set():
                self._set_status("Pysäytetty.")
                return
            if not rows:
                self._set_status("Ei löytynyt tuloksia.")
                return

            self._set_status("Tallennetaan tulokset…")
            out_dir = self._finalize_and_save(rows, emails)
            self.last_out_dir = out_dir
            self.last_emails = emails

            messagebox.showinfo(
                "Valmis",
                f"Valmis!\n\nKansio:\n{out_dir}\n\nRivejä: {len(rows)}\nSähköposteja: {len(emails)}"
            )
            self._set_status("Valmis!")
        except Exception as e:
            self._ui_log(f"VIRHE: {e}")
            messagebox.showerror("Virhe", f"Tuli virhe:\n\n{e}")


if __name__ == "__main__":
    App().mainloop()
