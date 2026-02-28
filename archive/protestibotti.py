# app.py
# LeadForge FI (Finnish B2B Email Finder) - Mass-market EXE MVP
#
# Features (MVP):
# - Modes: PDF -> YTJ, Clipboard -> Finder -> YTJ
# - KL-friendly name+location extraction with optional "Asunto Oy"
# - Name -> Y-tunnus -> Email, with location boost + optional "require location if present"
# - Email from YTJ (click "Näytä"), fallback to company website (YTJ website link) + contact pages
# - Outputs only on finalize: emails.docx, results.xlsx, results.csv
# - Autosave + Graceful stop (saves partial results)
# - Resume: rerun only failed rows from previous results.xlsx
# - License: Demo (no/invalid key) limited per run; Pro unlimited (offline check)
# - Safe/Fast mode toggle
#
# NOTE: No permanent log folder created unless results exist.
#       Temporary work happens in system temp and is deleted if no results.

import os
import re
import sys
import time
import csv
import hmac
import base64
import shutil
import tempfile
import threading
import subprocess
from dataclasses import dataclass
from difflib import SequenceMatcher
from urllib.parse import urljoin, urlparse
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import PyPDF2
from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD  # type: ignore
    HAS_DND = True
except Exception:
    HAS_DND = False

# =========================
#   BUILD / PRODUCT INFO
# =========================
APP_NAME = "LeadForge FI"
APP_BUILD = "2026-02-28_mvp_massmarket_license_resume_fallback"
WINDOW_TITLE = f"{APP_NAME}"

# =========================
#   LICENSE (OFFLINE MVP)
# =========================
LICENSE_SECRET = b"LeadForgeFI_MVP_secret_change_this_before_big_launch"
LICENSE_FILE = "license.key"
DEMO_LIMIT_PER_RUN = 20  # demo max companies per run (names or yts)
LICENSE_GRACE_DAYS = 30  # reserved for later expansion

KEY_RE = re.compile(r"^LF-[A-Z0-9]{4}-[A-Z0-9]{4}-[A-Z0-9]{4}-[A-Z0-9]{2}$", re.I)

def exe_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def license_path() -> str:
    return os.path.join(exe_dir(), LICENSE_FILE)

def _sig2(payload: str) -> str:
    mac = hmac.new(LICENSE_SECRET, payload.encode("utf-8"), digestmod="sha256").digest()
    return base64.b32encode(mac)[:2].decode("ascii")

def validate_license_key(key: str) -> bool:
    if not key:
        return False
    key = key.strip().upper()
    if not KEY_RE.match(key):
        return False
    parts = key.split("-")
    payload = "-".join(parts[:-1])  # LF-XXXX-XXXX-XXXX
    cc = parts[-1]
    return _sig2(payload) == cc

def read_saved_license() -> str:
    p = license_path()
    try:
        if os.path.exists(p):
            with open(p, "r", encoding="utf-8") as f:
                return (f.read() or "").strip()
    except Exception:
        pass
    return ""

def save_license(key: str) -> bool:
    try:
        with open(license_path(), "w", encoding="utf-8") as f:
            f.write((key or "").strip())
        return True
    except Exception:
        return False

# =========================
#   TUNING
# =========================
YTJ_PAGE_LOAD_TIMEOUT = 18
YTJ_RETRY_READS = 5
YTJ_RETRY_SLEEP = 0.12
YTJ_NAYTA_PASSES = 2

NAME_SEARCH_TIMEOUT = 10.0
NAME_SEARCH_SLEEP = 0.18

SAFE_SLEEP = 0.12
FAST_SLEEP = 0.02

WEB_TIMEOUT = 8
WEB_MAX_PAGES = 3  # homepage + up to 2 contact-ish pages

# =========================
#   REGEX
# =========================
YT_RE = re.compile(r"\b\d{7}-\d\b|\b\d{8}\b")
EMAIL_RE = re.compile(r"[A-Za-z0-9_.+-]+@[A-Za-z0-9-]+\.[A-Za-z0-9-.]+")
EMAIL_A_RE = re.compile(r"[A-Za-z0-9_.+-]+\s*\(a\)\s*[A-Za-z0-9-]+\.[A-Za-z0-9-.]+", re.I)
DATE_RE = re.compile(r"\b\d{1,2}\.\d{1,2}\.\d{4}\b")

STRICT_FORMS_RE = re.compile(
    r"\b(oy|ab|ky|tmi|oyj|osakeyhtiö|kommandiittiyhtiö|toiminimi|as\.|ltd|llc|inc|gmbh)\b",
    re.I,
)
ASUNTO_OY_RE = re.compile(r"\basunto\s+oy\b", re.I)

YTJ_COMPANY_URL = "https://tietopalvelu.ytj.fi/yritys/{}"
YTJ_SEARCH_URLS = ["https://tietopalvelu.ytj.fi/haku", "https://tietopalvelu.ytj.fi/"]

# =========================
#   SCROLLABLE UI WRAPPER
# =========================
class ScrollableFrame(ttk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)

        self.canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.inner = ttk.Frame(self.canvas)
        self._win = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _on_inner_configure(self, _event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self._win, width=event.width)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


BaseTk = TkinterDnD.Tk if HAS_DND else tk.Tk

# =========================
#   OUTPUT DIRS (LAZY)
# =========================
def base_output_dir() -> str:
    base = exe_dir()
    try:
        test = os.path.join(base, "_write_test.tmp")
        with open(test, "w", encoding="utf-8") as f:
            f.write("ok")
        os.remove(test)
        return os.path.join(base, "LeadForgeFI")
    except Exception:
        home = os.path.expanduser("~")
        docs = os.path.join(home, "Documents")
        return os.path.join(docs, "LeadForgeFI")

def create_final_run_dir() -> str:
    root = base_output_dir()
    date_folder = time.strftime("%Y-%m-%d")
    run_folder = "run_" + time.strftime("%H-%M-%S")
    out = os.path.join(root, date_folder, run_folder)
    os.makedirs(out, exist_ok=True)
    return out

def open_path_in_os(path: str):
    if not path:
        return
    try:
        if os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", path], check=False)
        else:
            subprocess.run(["xdg-open", path], check=False)
    except Exception:
        pass

# =========================
#   NORMALIZERS
# =========================
def normalize_yt(yt: str):
    yt = (yt or "").strip().replace(" ", "")
    if re.fullmatch(r"\d{7}-\d", yt):
        return yt
    if re.fullmatch(r"\d{8}", yt):
        return yt[:7] + "-" + yt[7]
    return None

def pick_email_from_text(text: str) -> str:
    if not text:
        return ""
    m = EMAIL_RE.search(text)
    if m:
        return m.group(0).strip().replace(" ", "")
    m2 = EMAIL_A_RE.search(text)
    if m2:
        return m2.group(0).replace(" ", "").replace("(a)", "@")
    return ""

def split_lines(text: str):
    if not text:
        return []
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return [ln.strip() for ln in text.split("\n") if ln.strip()]

def _looks_like_location(line: str) -> bool:
    if not line:
        return False
    s = re.sub(r"\s{2,}", " ", line).strip()
    if len(s) < 2 or len(s) > 32:
        return False
    low = s.lower()
    if "€" in s or DATE_RE.search(s):
        return False
    if YT_RE.search(s):
        return False
    if STRICT_FORMS_RE.search(s) or ASUNTO_OY_RE.search(s):
        return False
    if any(ch.isdigit() for ch in s):
        return False
    bad = ["yritys", "sijainti", "summa", "häiriöpäivä", "tyyppi", "lähde", "velkoja", "alue", "julkaisupäivä"]
    if any(b in low for b in bad):
        return False
    if not any(ch.isalpha() for ch in s):
        return False
    return True

def clean_company_name(name: str) -> str:
    n = (name or "").strip()
    n = re.sub(r"\s{2,}", " ", n)
    n = re.split(r"\s+[|•·]\s+|\s+-\s+|\s+–\s+", n)[0].strip()
    n = re.sub(r"\s*\(.*?\)\s*$", "", n).strip()
    return n

def extract_names_with_locations(text: str, strict: bool, allow_asunto: bool, max_names: int):
    lines = split_lines(text)
    out = []
    seen = set()

    bad_contains = [
        "näytä lisää", "kirjaudu", "tilaa", "tilaajille",
        "€", "y-tunnus", "ytunnus", "sähköposti", "puhelin",
        "www.", "http", "kauppalehti", "protestilista",
        "viimeisimmät protestit", "häiriöpäivä", "velkomustuomiot",
        "sijainti", "summa", "lähde", "tyyppi",
    ]

    def _is_name_candidate(ln: str) -> bool:
        if not ln or len(ln) < 3:
            return False
        if YT_RE.search(ln):
            return False
        low = ln.lower()
        if any(b in low for b in bad_contains):
            return False
        if sum(ch.isdigit() for ch in ln) >= 3:
            return False
        if not any(ch.isalpha() for ch in ln):
            return False
        nm = clean_company_name(ln)
        if len(nm) > 90:
            return False
        if strict:
            if STRICT_FORMS_RE.search(nm):
                return True
            if allow_asunto and ASUNTO_OY_RE.search(nm):
                return True
            return False
        return True

    for i, ln in enumerate(lines):
        if len(out) >= max_names:
            break
        if not _is_name_candidate(ln):
            continue

        name = clean_company_name(ln)
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)

        loc = ""
        for j in range(i + 1, min(i + 5, len(lines))):
            cand = lines[j]
            if _looks_like_location(cand):
                loc = re.sub(r"\s{2,}", " ", cand).strip()
                break

        out.append({"name": name, "location_hint": loc})
    return out

# =========================
#   OUTPUT WRITERS
# =========================
def save_emails_docx(out_dir: str, emails: list[str]):
    path = os.path.join(out_dir, "emails.docx")
    doc = Document()
    for e in emails:
        if e:
            doc.add_paragraph(e)
    doc.save(path)
    return path

def _autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 500) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(60, max(12, max_len + 2))

def save_results_xlsx(out_dir: str, rows: list[dict], filename="results.xlsx"):
    path = os.path.join(out_dir, filename)
    wb = Workbook()

    headers = ["Name", "Y-tunnus", "Email", "Status", "Source", "Notes", "LocationHint", "Website", "OtherEmails"]
    header_font = Font(bold=True)

    ws = wb.active
    ws.title = "Results"
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="left")

    for r in rows:
        ws.append([
            r.get("name", ""),
            r.get("yt", ""),
            r.get("email", ""),
            r.get("status", ""),
            r.get("source", ""),
            r.get("notes", ""),
            r.get("location_hint", ""),
            r.get("website", ""),
            r.get("other_emails", ""),
        ])
    _autosize_columns(ws)

    ws2 = wb.create_sheet("Found Only")
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws2.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="left")

    for r in rows:
        if (r.get("email") or "").strip():
            ws2.append([
                r.get("name", ""),
                r.get("yt", ""),
                r.get("email", ""),
                r.get("status", ""),
                r.get("source", ""),
                r.get("notes", ""),
                r.get("location_hint", ""),
                r.get("website", ""),
                r.get("other_emails", ""),
            ])
    _autosize_columns(ws2)

    ws3 = wb.create_sheet("Not Found")
    ws3.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws3.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="left")
    for r in rows:
        if not (r.get("email") or "").strip():
            ws3.append([
                r.get("name", ""),
                r.get("yt", ""),
                r.get("email", ""),
                r.get("status", ""),
                r.get("source", ""),
                r.get("notes", ""),
                r.get("location_hint", ""),
                r.get("website", ""),
                r.get("other_emails", ""),
            ])
    _autosize_columns(ws3)

    wb.save(path)
    return path

def save_results_csv(out_dir: str, rows: list[dict], filename="results.csv"):
    path = os.path.join(out_dir, filename)
    headers = ["Name","Y-tunnus","Email","Status","Source","Notes","LocationHint","Website","OtherEmails"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(headers)
        for r in rows:
            w.writerow([
                r.get("name",""),
                r.get("yt",""),
                r.get("email",""),
                r.get("status",""),
                r.get("source",""),
                r.get("notes",""),
                r.get("location_hint",""),
                r.get("website",""),
                r.get("other_emails",""),
            ])
    return path

# =========================
#   PDF -> YTs
# =========================
def extract_ytunnukset_from_pdf(pdf_path: str):
    yt_set = set()
    with open(pdf_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            text = page.extract_text() or ""
            for m in YT_RE.findall(text):
                n = normalize_yt(m)
                if n:
                    yt_set.add(n)
    return sorted(yt_set)

# =========================
#   WEB FALLBACK (urllib)
# =========================
def _fetch_url(url: str, timeout: int = WEB_TIMEOUT) -> str:
    if not url:
        return ""
    try:
        req = Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) LeadForgeFI/1.0",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            },
        )
        with urlopen(req, timeout=timeout) as r:
            data = r.read()
        try:
            return data.decode("utf-8", errors="ignore")
        except Exception:
            return data.decode("latin-1", errors="ignore")
    except (HTTPError, URLError, TimeoutError):
        return ""
    except Exception:
        return ""

def _extract_emails_from_html(html: str) -> list[str]:
    if not html:
        return []
    found = set()
    for e in EMAIL_RE.findall(html):
        e2 = e.strip().lower()
        if e2:
            found.add(e2)
    for e in EMAIL_A_RE.findall(html):
        e2 = e.replace(" ", "").replace("(a)", "@").lower()
        if e2:
            found.add(e2)
    return sorted(found)

def _extract_candidate_contact_links(base_url: str, html: str) -> list[str]:
    if not html or not base_url:
        return []
    hrefs = re.findall(r'href\s*=\s*["\']([^"\']+)["\']', html, flags=re.I)
    candidates = []
    for h in hrefs:
        h = h.strip()
        if not h or h.startswith("#") or h.startswith("mailto:") or h.startswith("tel:"):
            continue
        low = h.lower()
        if any(x in low for x in ["yhteys", "yhteystied", "contact", "kontakt", "asiakaspalvelu"]):
            absu = urljoin(base_url, h)
            candidates.append(absu)
    uniq = []
    seen = set()
    base_host = urlparse(base_url).netloc.lower()
    for u in candidates:
        if u in seen:
            continue
        seen.add(u)
        if urlparse(u).netloc.lower() == base_host:
            uniq.append(u)
    return uniq[:2]

def rank_emails(emails: list[str], website_url: str = "") -> tuple[str, str]:
    if not emails:
        return "", ""
    host = urlparse(website_url).netloc.lower().replace("www.", "")
    def score(e: str) -> float:
        el = e.lower()
        s = 0.0
        if "noreply" in el or "no-reply" in el or "donotreply" in el:
            s -= 100
        if host and host in el:
            s += 25
        if any(x in el for x in ["info@", "contact@", "asiakaspalvelu@", "sales@", "myynti@"]):
            s -= 5
        local = el.split("@", 1)[0]
        if len(local) >= 8:
            s += 2
        return s
    emails2 = sorted(set([e.lower().strip() for e in emails if e.strip()]))
    best = max(emails2, key=score)
    others = [e for e in emails2 if e != best]
    return best, ";".join(others)

def website_fallback_find_email(website_url: str) -> tuple[str, str]:
    if not website_url:
        return "", ""
    if not website_url.startswith("http"):
        website_url = "https://" + website_url
    html = _fetch_url(website_url)
    emails = _extract_emails_from_html(html)
    if emails:
        return rank_emails(emails, website_url)
    links = _extract_candidate_contact_links(website_url, html)
    for u in links[:max(0, WEB_MAX_PAGES-1)]:
        html2 = _fetch_url(u)
        emails2 = _extract_emails_from_html(html2)
        if emails2:
            return rank_emails(emails2, website_url)
    return "", ""

# =========================
#   SELENIUM (YTJ)
# =========================
def start_new_driver(headless: bool = False):
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    if headless:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1600,1000")

    driver_path = ChromeDriverManager().install()
    drv = webdriver.Chrome(service=Service(driver_path), options=options)
    drv.set_page_load_timeout(YTJ_PAGE_LOAD_TIMEOUT)
    return drv

def safe_click(driver, elem) -> bool:
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
        time.sleep(0.01)
        try:
            elem.click()
        except Exception:
            driver.execute_script("arguments[0].click();", elem)
        return True
    except Exception:
        return False

def try_accept_cookies(driver):
    texts = ["Hyväksy", "Hyväksy kaikki", "Salli kaikki", "Accept", "Accept all", "I agree", "OK", "Selvä"]
    for _ in range(2):
        for e in driver.find_elements(By.XPATH, "//button|//a|//*[@role='button']"):
            try:
                t = (e.text or "").strip()
                if not t:
                    continue
                if any(x.lower() in t.lower() for x in texts):
                    if e.is_displayed() and e.is_enabled():
                        safe_click(driver, e)
                        time.sleep(0.15)
                        break
            except Exception:
                continue

def wait_ytj_loaded(driver):
    WebDriverWait(driver, YTJ_PAGE_LOAD_TIMEOUT).until(
        EC.presence_of_element_located((By.TAG_NAME, "body"))
    )

def extract_email_from_ytj(driver):
    try:
        for a in driver.find_elements(By.TAG_NAME, "a"):
            href = (a.get_attribute("href") or "")
            if href.lower().startswith("mailto:"):
                return href.split(":", 1)[1].strip()
    except Exception:
        pass

    try:
        cells = driver.find_elements(By.XPATH, "//tr//*[self::td or self::th][contains(normalize-space(.), 'Sähköposti')]")
        for c in cells:
            try:
                tr = c.find_element(By.XPATH, "ancestor::tr[1]")
                email = pick_email_from_text(tr.text or "")
                if email:
                    return email
            except Exception:
                continue
    except Exception:
        pass

    try:
        return pick_email_from_text(driver.find_element(By.TAG_NAME, "body").text or "")
    except Exception:
        return ""

def extract_website_from_ytj(driver) -> str:
    try:
        for a in driver.find_elements(By.TAG_NAME, "a"):
            href = (a.get_attribute("href") or "").strip()
            if not href:
                continue
            low = href.lower()
            if low.startswith("mailto:") or low.startswith("tel:"):
                continue
            if "tietopalvelu.ytj.fi" in low:
                continue
            if low.startswith("http://") or low.startswith("https://"):
                return href
    except Exception:
        pass
    return ""

def click_all_nayta_ytj(driver):
    for _ in range(YTJ_NAYTA_PASSES):
        clicked = False
        for b in driver.find_elements(By.TAG_NAME, "button"):
            try:
                if (b.text or "").strip().lower() == "näytä" and b.is_displayed() and b.is_enabled():
                    safe_click(driver, b)
                    clicked = True
                    time.sleep(0.06)
            except Exception:
                continue
        for a in driver.find_elements(By.TAG_NAME, "a"):
            try:
                if (a.text or "").strip().lower() == "näytä" and a.is_displayed():
                    safe_click(driver, a)
                    clicked = True
                    time.sleep(0.06)
            except Exception:
                continue
        if not clicked:
            break

def fetch_email_by_yt(driver, yt: str, stop_flag: threading.Event, do_web_fallback: bool = True):
    if stop_flag.is_set():
        return "", "", "", "stopped", "stop requested"

    try:
        driver.get(YTJ_COMPANY_URL.format(yt))
    except TimeoutException:
        pass
    try:
        wait_ytj_loaded(driver)
    except Exception:
        pass
    try_accept_cookies(driver)

    for _ in range(2):
        if stop_flag.is_set():
            return "", "", "", "stopped", "stop requested"
        email = extract_email_from_ytj(driver)
        if email:
            return email, "", "", "ok", "email from YTJ"
        time.sleep(YTJ_RETRY_SLEEP)

    click_all_nayta_ytj(driver)
    for _ in range(YTJ_RETRY_READS):
        if stop_flag.is_set():
            return "", "", "", "stopped", "stop requested"
        email = extract_email_from_ytj(driver)
        if email:
            return email, "", "", "ok", "email after Näytä"
        time.sleep(YTJ_RETRY_SLEEP)

    website = extract_website_from_ytj(driver)
    if do_web_fallback and website and not stop_flag.is_set():
        best, others = website_fallback_find_email(website)
        if best:
            return best, website, others, "website_fallback_ok", "email from website"
        return "", website, "", "website_no_email", "no email on YTJ or website"
    return "", website, "", "yt_found_no_email", "no email on YTJ"

# =========================
#   YTJ SEARCH
# =========================
def _attr(el, name: str) -> str:
    try:
        return (el.get_attribute(name) or "").strip()
    except Exception:
        return ""

def find_ytj_company_search_input(driver):
    candidates = []
    try:
        inputs = driver.find_elements(By.XPATH, "//input")
    except Exception:
        inputs = []

    for inp in inputs:
        try:
            if not inp.is_displayed() or not inp.is_enabled():
                continue
            itype = (_attr(inp, "type") or "").lower()
            if itype in ("hidden", "password", "checkbox", "radio", "submit", "button", "file"):
                continue

            ph = _attr(inp, "placeholder").lower()
            al = _attr(inp, "aria-label").lower()
            nm = _attr(inp, "name").lower()
            iid = _attr(inp, "id").lower()
            text = " ".join([ph, al, nm, iid]).strip()

            if "y-tunnus" in text and ("yritys" not in text and "toiminimi" not in text and "nimi" not in text):
                continue

            score = 0
            if "yritys" in text:
                score += 50
            if "toiminimi" in text:
                score += 35
            if "nimi" in text:
                score += 25
            if itype == "search":
                score += 15
            if "hae" in text:
                score += 10
            if "y-tunnus" in text:
                score -= 5

            if score <= 0:
                continue
            candidates.append((score, inp))
        except Exception:
            continue

    if not candidates:
        try:
            for inp in driver.find_elements(By.XPATH, "//input[@type='search']"):
                if inp.is_displayed() and inp.is_enabled():
                    return inp
        except Exception:
            pass
        return None

    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]

def ensure_ytj_search_ready(driver) -> bool:
    if find_ytj_company_search_input(driver):
        return True
    for url in YTJ_SEARCH_URLS:
        try:
            driver.get(url)
            wait_ytj_loaded(driver)
            try_accept_cookies(driver)
            if find_ytj_company_search_input(driver):
                return True
        except Exception:
            continue
    return False

def extract_yt_from_text_anywhere(txt: str) -> str:
    if not txt:
        return ""
    for m in YT_RE.findall(txt):
        n = normalize_yt(m)
        if n:
            return n
    return ""

def score_result(name_query: str, card_text: str, location_hint: str = "", use_location_boost: bool = True) -> float:
    txt = (card_text or "").strip()
    ratio = SequenceMatcher(None, (name_query or "").lower(), txt.lower()).ratio()
    score = ratio * 100.0
    if extract_yt_from_text_anywhere(txt):
        score += 20.0
    if use_location_boost and location_hint:
        if location_hint.lower() in txt.lower():
            score += 35.0
    if ASUNTO_OY_RE.search(name_query) and ASUNTO_OY_RE.search(txt):
        score += 10.0
    return score

def open_href_in_new_tab_and_extract_yt(driver, href: str) -> str:
    if not href:
        return ""
    base_handle = driver.current_window_handle
    before_handles = set(driver.window_handles)

    try:
        driver.execute_script("window.open(arguments[0], '_blank');", href)
        t0 = time.time()
        new_handle = None
        while time.time() - t0 < 5.0:
            diff = [h for h in driver.window_handles if h not in before_handles]
            if diff:
                new_handle = diff[0]
                break
            time.sleep(0.05)
        if not new_handle:
            return ""

        driver.switch_to.window(new_handle)
        try:
            wait_ytj_loaded(driver)
        except Exception:
            pass
        try_accept_cookies(driver)

        try:
            body = driver.find_element(By.TAG_NAME, "body").text or ""
        except Exception:
            body = ""
        yt = extract_yt_from_text_anywhere(body)

        try:
            driver.close()
        except Exception:
            pass
        try:
            driver.switch_to.window(base_handle)
        except Exception:
            for h in driver.window_handles:
                try:
                    driver.switch_to.window(h)
                    break
                except Exception:
                    continue
        return yt
    except Exception:
        try:
            for h in driver.window_handles:
                if h != base_handle:
                    driver.switch_to.window(h)
                    driver.close()
        except Exception:
            pass
        try:
            driver.switch_to.window(base_handle)
        except Exception:
            pass
        return ""

def ytj_name_to_yt(driver, name: str, stop_flag: threading.Event, location_hint: str = "", use_location_boost: bool = True):
    if stop_flag.is_set():
        return ""
    if not ensure_ytj_search_ready(driver):
        return ""
    if stop_flag.is_set():
        return ""

    inp = find_ytj_company_search_input(driver)
    if not inp:
        return ""

    try:
        try:
            inp.clear()
        except Exception:
            pass
        inp.send_keys(name)
        inp.send_keys(u"\ue007")  # ENTER
    except Exception:
        return ""

    best_href = ""
    best_score = -1.0
    t0 = time.time()

    while time.time() - t0 < NAME_SEARCH_TIMEOUT:
        if stop_flag.is_set():
            return ""

        candidate_links = []
        for xp in ("//a[contains(@href,'/yritys/')]", "//a[contains(@href,'yritys')]"):
            try:
                candidate_links.extend(driver.find_elements(By.XPATH, xp))
            except Exception:
                pass

        checked = 0
        for a in candidate_links:
            if checked >= 30:
                break
            try:
                if not a.is_displayed():
                    continue
                href = (a.get_attribute("href") or "")
                if not href or "tietopalvelu.ytj.fi" not in href:
                    continue

                try:
                    card = a.find_element(By.XPATH, "ancestor::*[self::li or self::div or self::article][1]")
                    card_text = (card.text or "")
                except Exception:
                    card_text = (a.text or "")

                s = score_result(name, card_text, location_hint=location_hint, use_location_boost=use_location_boost)
                checked += 1
                if s > best_score:
                    best_score = s
                    best_href = href
            except Exception:
                continue

        if best_href:
            break
        time.sleep(NAME_SEARCH_SLEEP)

    if not best_href:
        return ""
    return open_href_in_new_tab_and_extract_yt(driver, best_href)

# =========================
#   RESUME FROM XLSX
# =========================
def load_failed_rows_from_results_xlsx(path: str) -> list[dict]:
    wb = load_workbook(path)
    ws = wb["Results"] if "Results" in wb.sheetnames else wb.active
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[str(v).strip()] = c

    def get(row, name, default=""):
        col = headers.get(name)
        if not col:
            return default
        v = ws.cell(row=row, column=col).value
        return "" if v is None else str(v)

    out = []
    for r in range(2, ws.max_row + 1):
        email = get(r, "Email")
        status = get(r, "Status")
        name = get(r, "Name")
        yt = get(r, "Y-tunnus")
        loc = get(r, "LocationHint")
        src = get(r, "Source")
        notes = get(r, "Notes")
        website = get(r, "Website")
        other = get(r, "OtherEmails")

        if email.strip():
            continue
        if not name.strip() and not yt.strip():
            continue
        out.append({
            "name": name.strip(),
            "yt": yt.strip(),
            "email": "",
            "status": status.strip() or "resume",
            "source": (src.strip() or "resume"),
            "notes": (notes.strip() or "resume rerun"),
            "location_hint": loc.strip(),
            "website": website.strip(),
            "other_emails": other.strip(),
        })
    return out

# =========================
#   TEMP WORKSPACE
# =========================
@dataclass
class WorkContext:
    temp_dir: str
    rows: list
    emails: list
    started_at: float

def create_work_context() -> WorkContext:
    td = tempfile.mkdtemp(prefix="LeadForgeFI_")
    return WorkContext(temp_dir=td, rows=[], emails=[], started_at=time.time())

def cleanup_temp_dir(temp_dir: str):
    try:
        if temp_dir and os.path.isdir(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception:
        pass

def finalize_outputs(ctx: WorkContext) -> str:
    if not ctx.rows:
        cleanup_temp_dir(ctx.temp_dir)
        return ""
    final_dir = create_final_run_dir()

    save_results_xlsx(ctx.temp_dir, ctx.rows)
    save_results_csv(ctx.temp_dir, ctx.rows)
    save_emails_docx(ctx.temp_dir, ctx.emails)

    for fn in ("results.xlsx", "results.csv", "emails.docx"):
        src = os.path.join(ctx.temp_dir, fn)
        if os.path.exists(src):
            shutil.copy2(src, os.path.join(final_dir, fn))

    cleanup_temp_dir(ctx.temp_dir)
    return final_dir

# =========================
#   PIPELINES
# =========================
def pipeline_pdf(ctx: WorkContext, pdf_path: str, status_cb, progress_cb, stop_flag: threading.Event,
                 headless: bool, safe_mode: bool, do_web_fallback: bool, demo_mode: bool):
    status_cb("Luetaan PDF ja kerätään Y-tunnukset…")
    yts = extract_ytunnukset_from_pdf(pdf_path)
    if not yts:
        status_cb("PDF: ei löytynyt Y-tunnuksia.")
        return

    if demo_mode:
        yts = yts[:DEMO_LIMIT_PER_RUN]

    status_cb(f"PDF: löytyi {len(yts)} Y-tunnusta. Haetaan emailit YTJ:stä…")
    driver = start_new_driver(headless=headless)
    sleep_each = SAFE_SLEEP if safe_mode else FAST_SLEEP

    try:
        progress_cb(0, max(1, len(yts)))
        cache = {}
        for i, yt in enumerate(yts, start=1):
            if stop_flag.is_set():
                status_cb("Pysäytetty — tallennetaan tähän asti…")
                break
            progress_cb(i - 1, len(yts))
            status_cb(f"YTJ: {i}/{len(yts)} {yt}")

            if yt in cache:
                email, website, other, st, notes = cache[yt]
            else:
                email, website, other, st, notes = fetch_email_by_yt(driver, yt, stop_flag, do_web_fallback=do_web_fallback)
                cache[yt] = (email, website, other, st, notes)

            ctx.rows.append({
                "name": "",
                "yt": yt,
                "email": email,
                "status": st,
                "source": "pdf->ytj",
                "notes": notes,
                "location_hint": "",
                "website": website,
                "other_emails": other,
            })
            time.sleep(sleep_each)

        progress_cb(len(yts), max(1, len(yts)))
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    ctx.emails = sorted({(r.get("email") or "").lower(): (r.get("email") or "") for r in ctx.rows if (r.get("email") or "").strip()}.values())

def pipeline_clipboard(ctx: WorkContext, text: str, strict: bool, allow_asunto: bool, max_names: int,
                       use_location_boost: bool, require_location_if_present: bool,
                       status_cb, progress_cb, stop_flag: threading.Event,
                       headless: bool, safe_mode: bool, do_web_fallback: bool, demo_mode: bool):
    status_cb("Clipboard: poimitaan sähköpostit ja yritysnimet…")

    direct_emails = sorted({e.strip().lower() for e in EMAIL_RE.findall(text or "") if e.strip()})
    name_items = extract_names_with_locations(text, strict=strict, allow_asunto=allow_asunto, max_names=max_names)

    for em in direct_emails:
        ctx.rows.append({
            "name": "",
            "yt": "",
            "email": em,
            "status": "ok",
            "source": "clipboard",
            "notes": "email found in pasted text",
            "location_hint": "",
            "website": "",
            "other_emails": "",
        })

    if not name_items and not direct_emails:
        status_cb("Clipboard: en löytänyt mitään (email / nimi).")
        return

    if demo_mode:
        name_items = name_items[:DEMO_LIMIT_PER_RUN]

    status_cb("Käynnistetään YTJ-haku…")
    driver = start_new_driver(headless=headless)
    sleep_each = SAFE_SLEEP if safe_mode else FAST_SLEEP

    yt_cache: dict[str, tuple] = {}
    name_cache: dict[str, str] = {}

    tasks: list[dict] = []
    seen = set()
    for it in name_items:
        nm = (it.get("name") or "").strip()
        if not nm:
            continue
        key = nm.lower()
        if key in seen:
            continue
        seen.add(key)
        tasks.append({
            "name": nm,
            "location_hint": (it.get("location_hint") or "").strip(),
            "yt": "",
            "email": "",
            "status": "",
            "source": "clipboard->ytj",
            "notes": "name->yt->email",
            "website": "",
            "other_emails": "",
        })

    try:
        ensure_ytj_search_ready(driver)

        progress_cb(0, max(1, len(tasks)))
        for idx, r in enumerate(tasks, start=1):
            if stop_flag.is_set():
                status_cb("Pysäytetty — tallennetaan tähän asti…")
                break
            progress_cb(idx - 1, len(tasks))

            name = r["name"]
            loc = r.get("location_hint", "")

            status_cb(f"YTJ yrityshaku: {idx}/{len(tasks)} {name}" + (f" ({loc})" if loc else ""))

            cache_key = name.lower() + ("||" + loc.lower() if (use_location_boost and loc) else "")
            yt = name_cache.get(cache_key)
            if yt is None:
                yt = ytj_name_to_yt(driver, name, stop_flag, location_hint=loc, use_location_boost=use_location_boost)
                name_cache[cache_key] = yt

            if require_location_if_present and loc and not yt:
                status_cb(f"Fallback ilman sijaintia: {name}")
                yt = name_cache.get(name.lower())
                if yt is None:
                    yt = ytj_name_to_yt(driver, name, stop_flag, location_hint="", use_location_boost=False)
                    name_cache[name.lower()] = yt

            if not yt:
                r["status"] = "yt_not_found"
                r["notes"] = "yt not found"
                ctx.rows.append(r)
                time.sleep(sleep_each)
                continue

            r["yt"] = yt
            status_cb(f"YTJ email: {idx}/{len(tasks)} {yt}")

            if yt in yt_cache:
                email, website, other, st, notes = yt_cache[yt]
            else:
                email, website, other, st, notes = fetch_email_by_yt(driver, yt, stop_flag, do_web_fallback=do_web_fallback)
                yt_cache[yt] = (email, website, other, st, notes)

            r["email"] = email
            r["website"] = website
            r["other_emails"] = other
            r["status"] = st if st else ("ok" if email else "yt_found_no_email")
            r["notes"] = notes

            ctx.rows.append(r)
            time.sleep(sleep_each)

        progress_cb(len(tasks), max(1, len(tasks)))
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    ctx.emails = sorted({(rr.get("email") or "").lower(): (rr.get("email") or "") for rr in ctx.rows if (rr.get("email") or "").strip()}.values())

def pipeline_resume_failed(ctx: WorkContext, xlsx_path: str,
                           use_location_boost: bool, require_location_if_present: bool,
                           status_cb, progress_cb, stop_flag: threading.Event,
                           headless: bool, safe_mode: bool, do_web_fallback: bool, demo_mode: bool):
    status_cb("Ladataan Not Found rivit results.xlsx:stä…")
    tasks = load_failed_rows_from_results_xlsx(xlsx_path)
    if not tasks:
        status_cb("Ei löytynyt epäonnistuneita rivejä.")
        return

    if demo_mode:
        tasks = tasks[:DEMO_LIMIT_PER_RUN]

    status_cb(f"Resume: {len(tasks)} riviä. Käynnistetään YTJ-haku…")
    driver = start_new_driver(headless=headless)
    sleep_each = SAFE_SLEEP if safe_mode else FAST_SLEEP

    yt_cache: dict[str, tuple] = {}
    name_cache: dict[str, str] = {}

    try:
        ensure_ytj_search_ready(driver)
        progress_cb(0, max(1, len(tasks)))

        for idx, r in enumerate(tasks, start=1):
            if stop_flag.is_set():
                status_cb("Pysäytetty — tallennetaan tähän asti…")
                break
            progress_cb(idx - 1, len(tasks))

            name = (r.get("name") or "").strip()
            yt = (r.get("yt") or "").strip()
            loc = (r.get("location_hint") or "").strip()

            if yt:
                status_cb(f"YTJ email (resume): {idx}/{len(tasks)} {yt}")
            else:
                if not name:
                    continue
                status_cb(f"YTJ yrityshaku (resume): {idx}/{len(tasks)} {name}" + (f" ({loc})" if loc else ""))
                cache_key = name.lower() + ("||" + loc.lower() if (use_location_boost and loc) else "")
                yt = name_cache.get(cache_key)
                if yt is None:
                    yt = ytj_name_to_yt(driver, name, stop_flag, location_hint=loc, use_location_boost=use_location_boost)
                    name_cache[cache_key] = yt

                if require_location_if_present and loc and not yt:
                    yt = name_cache.get(name.lower())
                    if yt is None:
                        yt = ytj_name_to_yt(driver, name, stop_flag, location_hint="", use_location_boost=False)
                        name_cache[name.lower()] = yt

                if not yt:
                    r["status"] = "yt_not_found"
                    r["notes"] = "yt not found (resume)"
                    ctx.rows.append(r)
                    time.sleep(sleep_each)
                    continue
                r["yt"] = yt

            if yt in yt_cache:
                email, website, other, st, notes = yt_cache[yt]
            else:
                email, website, other, st, notes = fetch_email_by_yt(driver, yt, stop_flag, do_web_fallback=do_web_fallback)
                yt_cache[yt] = (email, website, other, st, notes)

            r["email"] = email
            r["website"] = website
            r["other_emails"] = other
            r["status"] = st if st else ("ok" if email else "yt_found_no_email")
            r["notes"] = notes + " | resume"

            ctx.rows.append(r)
            time.sleep(sleep_each)

        progress_cb(len(tasks), max(1, len(tasks)))
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    ctx.emails = sorted({(rr.get("email") or "").lower(): (rr.get("email") or "") for rr in ctx.rows if (rr.get("email") or "").strip()}.values())

# =========================
#   APP UI
# =========================
class App(BaseTk):
    def __init__(self):
        super().__init__()
        self.stop_flag = threading.Event()
        self.ctx: WorkContext | None = None
        self.last_emails: list[str] = []
        self.last_out_dir: str = ""
        self.license_key_var = tk.StringVar(value=read_saved_license())

        self.BG = "#ffffff"
        self.CARD = "#f6f8ff"
        self.BORDER = "#d7def7"
        self.TEXT = "#0b1b3a"
        self.MUTED = "#465a82"
        self.BLUE = "#1d4ed8"
        self.BLUE_H = "#1e40af"
        self.DANGER = "#b91c1c"
        self.DANGER_H = "#991b1b"
        self.GREY_BTN = "#64748b"
        self.GREY_BTN_H = "#475569"
        self.GREEN = "#047857"

        self.title(WINDOW_TITLE)
        self.geometry("1000x1000")
        self.configure(bg=self.BG)

        self._build_ui()
        self._refresh_license_ui()

    def _card(self, parent):
        return tk.Frame(parent, bg=self.CARD, highlightthickness=1, highlightbackground=self.BORDER)

    def _btn(self, parent, text, cmd, kind="blue"):
        if kind == "danger":
            bg, hover = self.DANGER, self.DANGER_H
        elif kind == "grey":
            bg, hover = self.GREY_BTN, self.GREY_BTN_H
        else:
            bg, hover = self.BLUE, self.BLUE_H

        b = tk.Button(
            parent, text=text, command=cmd,
            bg=bg, fg="#ffffff",
            activebackground=hover, activeforeground="#ffffff",
            relief="flat", padx=14, pady=10,
            font=("Segoe UI", 10, "bold")
        )
        b.bind("<Enter>", lambda e: b.configure(bg=hover))
        b.bind("<Leave>", lambda e: b.configure(bg=bg))
        return b

    def _ui_log(self, msg):
        ts = time.strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        try:
            self.listbox.insert(tk.END, line)
            self.listbox.yview_moveto(1.0)
        except Exception:
            pass

    def _set_status(self, s):
        self.status.config(text=s)
        self.update_idletasks()
        self._ui_log(s)

    def _set_progress(self, v, mx):
        self.progress["maximum"] = mx
        self.progress["value"] = v
        self.update_idletasks()

    def request_stop(self):
        self.stop_flag.set()
        self._set_status("Pysäytetään…")

    def _clear_stop(self):
        self.stop_flag.clear()

    def clear_clipboard_text(self):
        self.clip_text.delete("1.0", tk.END)
        self._ui_log("Clipboard tyhjennetty.")

    def open_out_folder(self):
        if not self.last_out_dir or not os.path.isdir(self.last_out_dir):
            messagebox.showinfo("Ei kansiota", "Aja ensin haku. Kansio syntyy vasta, kun tulokset tallennetaan.")
            return
        open_path_in_os(self.last_out_dir)

    def open_results_file(self):
        if not self.last_out_dir:
            messagebox.showinfo("Ei tiedostoa", "Aja ensin haku.")
            return
        p = os.path.join(self.last_out_dir, "results.xlsx")
        if os.path.exists(p):
            open_path_in_os(p)
        else:
            messagebox.showinfo("Ei tiedostoa", "results.xlsx ei löytynyt.")

    def copy_emails_to_clipboard(self):
        if not self.last_emails:
            messagebox.showinfo("Ei sähköposteja", "Aja haku ensin.")
            return
        txt = "\n".join(self.last_emails)
        try:
            self.clipboard_clear()
            self.clipboard_append(txt)
            self._ui_log(f"Kopioitu {len(self.last_emails)} sähköpostia.")
        except Exception:
            messagebox.showwarning("Kopiointi epäonnistui", "En saanut kopioitua leikepöydälle.")

    def copy_name_email_to_clipboard(self):
        if not self.ctx or not self.ctx.rows:
            messagebox.showinfo("Ei dataa", "Aja haku ensin.")
            return
        lines = []
        for r in self.ctx.rows:
            em = (r.get("email") or "").strip()
            nm = (r.get("name") or "").strip()
            if em:
                lines.append(f"{nm};{em}" if nm else f";{em}")
        if not lines:
            messagebox.showinfo("Ei löytynyt", "Ei löytynyt sähköposteja kopioitavaksi.")
            return
        txt = "\n".join(lines)
        try:
            self.clipboard_clear()
            self.clipboard_append(txt)
            self._ui_log(f"Kopioitu {len(lines)} riviä muodossa name;email.")
        except Exception:
            messagebox.showwarning("Kopiointi epäonnistui", "En saanut kopioitua leikepöydälle.")

    def _is_pro(self) -> bool:
        k = (self.license_key_var.get() or "").strip()
        return validate_license_key(k)

    def _refresh_license_ui(self):
        if self._is_pro():
            self.license_status.config(text="PRO aktiivinen", fg=self.GREEN)
        else:
            self.license_status.config(text=f"DEMO (max {DEMO_LIMIT_PER_RUN} / ajo)", fg=self.DANGER)

    def save_license_clicked(self):
        k = (self.license_key_var.get() or "").strip()
        if not k:
            messagebox.showwarning("Puuttuu", "Syötä lisenssiavain.")
            return
        if not validate_license_key(k):
            messagebox.showerror("Virhe", "Lisenssiavain ei kelpaa (muoto LF-XXXX-XXXX-XXXX-CC).")
            self._refresh_license_ui()
            return
        if save_license(k):
            messagebox.showinfo("OK", "Lisenssi tallennettu.")
        else:
            messagebox.showwarning("Virhe", "Lisenssin tallennus epäonnistui.")
        self._refresh_license_ui()

    def _finalize_and_save(self):
        if not self.ctx:
            return ""
        out_dir = finalize_outputs(self.ctx)
        self.last_out_dir = out_dir
        self.last_emails = self.ctx.emails or []
        return out_dir

    def _run_pdf(self, pdf_path: str):
        self.ctx = create_work_context()
        self.last_emails = []
        self.last_out_dir = ""
        demo_mode = not self._is_pro()
        try:
            self._set_status("Aloitetaan PDF ajo…")
            pipeline_pdf(
                self.ctx, pdf_path,
                self._set_status, self._set_progress,
                self.stop_flag,
                headless=bool(self.headless_var.get()),
                safe_mode=bool(self.safe_mode_var.get()),
                do_web_fallback=bool(self.web_fallback_var.get()),
                demo_mode=demo_mode,
            )
            if not self.ctx.rows:
                self._set_status("Ei tuloksia.")
                return

            self._set_status("Tallennetaan tulokset…")
            out_dir = self._finalize_and_save()
            if not out_dir:
                self._set_status("Ei tallennettavaa.")
                return

            msg = f"Valmis!\n\nKansio:\n{out_dir}\n\nRivejä: {len(self.ctx.rows)}\nSähköposteja: {len(self.ctx.emails)}"
            if demo_mode:
                msg += f"\n\n(DEMO: max {DEMO_LIMIT_PER_RUN} / ajo)"
            messagebox.showinfo("Valmis", msg)
            self._set_status("Valmis!")
        except Exception as e:
            self._ui_log(f"VIRHE: {e}")
            messagebox.showerror("Virhe", f"Tuli virhe:\n\n{e}")
        finally:
            if self.ctx and (not self.ctx.rows) and self.ctx.temp_dir:
                cleanup_temp_dir(self.ctx.temp_dir)

    def _run_clipboard(self, text: str):
        self.ctx = create_work_context()
        self.last_emails = []
        self.last_out_dir = ""
        demo_mode = not self._is_pro()
        try:
            self._set_status("Aloitetaan Clipboard ajo…")
            pipeline_clipboard(
                self.ctx, text,
                strict=bool(self.strict_var.get()),
                allow_asunto=bool(self.allow_asunto_var.get()),
                max_names=int(self.max_names_var.get() or 400),
                use_location_boost=bool(self.location_boost_var.get()),
                require_location_if_present=bool(self.require_location_var.get()),
                status_cb=self._set_status, progress_cb=self._set_progress,
                stop_flag=self.stop_flag,
                headless=bool(self.headless_var.get()),
                safe_mode=bool(self.safe_mode_var.get()),
                do_web_fallback=bool(self.web_fallback_var.get()),
                demo_mode=demo_mode,
            )
            if not self.ctx.rows:
                self._set_status("Ei tuloksia.")
                return

            self._set_status("Tallennetaan tulokset…")
            out_dir = self._finalize_and_save()
            if not out_dir:
                self._set_status("Ei tallennettavaa.")
                return

            msg = f"Valmis!\n\nKansio:\n{out_dir}\n\nRivejä: {len(self.ctx.rows)}\nSähköposteja: {len(self.ctx.emails)}"
            if demo_mode:
                msg += f"\n\n(DEMO: max {DEMO_LIMIT_PER_RUN} / ajo)"
            messagebox.showinfo("Valmis", msg)
            self._set_status("Valmis!")
        except Exception as e:
            self._ui_log(f"VIRHE: {e}")
            messagebox.showerror("Virhe", f"Tuli virhe:\n\n{e}")
        finally:
            if self.ctx and (not self.ctx.rows) and self.ctx.temp_dir:
                cleanup_temp_dir(self.ctx.temp_dir)

    def _run_resume(self, xlsx_path: str):
        self.ctx = create_work_context()
        self.last_emails = []
        self.last_out_dir = ""
        demo_mode = not self._is_pro()
        try:
            self._set_status("Aloitetaan Resume (vain epäonnistuneet)…")
            pipeline_resume_failed(
                self.ctx, xlsx_path,
                use_location_boost=bool(self.location_boost_var.get()),
                require_location_if_present=bool(self.require_location_var.get()),
                status_cb=self._set_status, progress_cb=self._set_progress,
                stop_flag=self.stop_flag,
                headless=bool(self.headless_var.get()),
                safe_mode=bool(self.safe_mode_var.get()),
                do_web_fallback=bool(self.web_fallback_var.get()),
                demo_mode=demo_mode,
            )

            if not self.ctx.rows:
                self._set_status("Ei tuloksia.")
                return

            self._set_status("Tallennetaan tulokset…")
            out_dir = self._finalize_and_save()
            if not out_dir:
                self._set_status("Ei tallennettavaa.")
                return

            msg = f"Valmis!\n\nKansio:\n{out_dir}\n\nRivejä: {len(self.ctx.rows)}\nSähköposteja: {len(self.ctx.emails)}"
            if demo_mode:
                msg += f"\n\n(DEMO: max {DEMO_LIMIT_PER_RUN} / ajo)"
            messagebox.showinfo("Valmis", msg)
            self._set_status("Valmis!")
        except Exception as e:
            self._ui_log(f"VIRHE: {e}")
            messagebox.showerror("Virhe", f"Tuli virhe:\n\n{e}")
        finally:
            if self.ctx and (not self.ctx.rows) and self.ctx.temp_dir:
                cleanup_temp_dir(self.ctx.temp_dir)

    def start_pdf_mode(self):
        self._clear_stop()
        path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if path:
            self.drop_var.set(f"PDF valittu: {path}")
            threading.Thread(target=self._run_pdf, args=(path,), daemon=True).start()

    def start_clipboard_mode(self):
        self._clear_stop()
        text = self.clip_text.get("1.0", tk.END).strip()
        if not text:
            messagebox.showwarning("Tyhjä", "Liitä ensin teksti (Ctrl+V).")
            return
        threading.Thread(target=self._run_clipboard, args=(text,), daemon=True).start()

    def start_resume_mode(self):
        self._clear_stop()
        path = filedialog.askopenfilename(filetypes=[("Excel results", "*.xlsx")])
        if not path:
            return
        threading.Thread(target=self._run_resume, args=(path,), daemon=True).start()

    def _on_drop_pdf(self, event):
        path = (event.data or "").strip()
        if path.startswith("{") and path.endswith("}"):
            path = path[1:-1]
        if path.lower().endswith(".pdf") and os.path.exists(path):
            self.drop_var.set(f"PDF valittu: {path}")
            self._clear_stop()
            threading.Thread(target=self._run_pdf, args=(path,), daemon=True).start()
        else:
            messagebox.showwarning("Ei PDF", "Pudotettu tiedosto ei ollut .pdf")

    def _build_ui(self):
        outer = ScrollableFrame(self)
        outer.pack(fill="both", expand=True)
        root = outer.inner

        header = tk.Frame(root, bg=self.BG)
        header.pack(fill="x", padx=16, pady=(16, 10))

        tk.Label(header, text=APP_NAME, bg=self.BG, fg=self.TEXT, font=("Segoe UI", 20, "bold")).pack(anchor="w")
        tk.Label(header, text=f"Build: {APP_BUILD}", bg=self.BG, fg=self.MUTED, font=("Segoe UI", 9)).pack(anchor="w")

        lic = self._card(root)
        lic.pack(fill="x", padx=16, pady=(0, 12))
        row = tk.Frame(lic, bg=self.CARD)
        row.pack(fill="x", padx=12, pady=12)

        tk.Label(row, text="Lisenssi:", bg=self.CARD, fg=self.TEXT, font=("Segoe UI", 10, "bold")).pack(side="left")
        ent = tk.Entry(row, textvariable=self.license_key_var, width=34, bg="#ffffff", fg=self.TEXT)
        ent.pack(side="left", padx=8)
        self._btn(row, "Tallenna", self.save_license_clicked, kind="grey").pack(side="left", padx=6)

        self.license_status = tk.Label(row, text="", bg=self.CARD, fg=self.MUTED, font=("Segoe UI", 10, "bold"))
        self.license_status.pack(side="right")

        actions = self._card(root)
        actions.pack(fill="x", padx=16, pady=(0, 12))

        row2 = tk.Frame(actions, bg=self.CARD)
        row2.pack(fill="x", padx=12, pady=(12, 8))

        self._btn(row2, "PDF → YTJ", self.start_pdf_mode).pack(side="left", padx=6)
        self._btn(row2, "Clipboard → Finder", self.start_clipboard_mode).pack(side="left", padx=6)
        self._btn(row2, "Resume (vain epäonnistuneet)", self.start_resume_mode, kind="grey").pack(side="left", padx=6)

        self._btn(row2, "Pysäytä", self.request_stop, kind="danger").pack(side="right", padx=6)

        row3 = tk.Frame(actions, bg=self.CARD)
        row3.pack(fill="x", padx=12, pady=(0, 12))

        self._btn(row3, "Avaa tuloskansio", self.open_out_folder, kind="grey").pack(side="left", padx=6)
        self._btn(row3, "Avaa results.xlsx", self.open_results_file, kind="grey").pack(side="left", padx=6)
        self._btn(row3, "Kopioi sähköpostit", self.copy_emails_to_clipboard, kind="grey").pack(side="right", padx=6)
        self._btn(row3, "Kopioi name;email", self.copy_name_email_to_clipboard, kind="grey").pack(side="right", padx=6)

        settings = self._card(root)
        settings.pack(fill="x", padx=16, pady=(0, 12))
        srow = tk.Frame(settings, bg=self.CARD)
        srow.pack(fill="x", padx=12, pady=12)

        self.safe_mode_var = tk.BooleanVar(value=True)
        self.web_fallback_var = tk.BooleanVar(value=True)
        self.headless_var = tk.BooleanVar(value=False)

        tk.Checkbutton(srow, text="Safe mode (suositus)", variable=self.safe_mode_var,
                       bg=self.CARD, fg=self.TEXT, selectcolor="#fff").pack(side="left")
        tk.Checkbutton(srow, text="Website fallback (jos YTJ ei anna emailia)", variable=self.web_fallback_var,
                       bg=self.CARD, fg=self.TEXT, selectcolor="#fff").pack(side="left", padx=(16,0))
        tk.Checkbutton(srow, text="Headless (nopeampi, vähemmän visuaalia)", variable=self.headless_var,
                       bg=self.CARD, fg=self.TEXT, selectcolor="#fff").pack(side="left", padx=(16,0))

        status_card = self._card(root)
        status_card.pack(fill="x", padx=16, pady=(0, 12))

        self.status = tk.Label(status_card, text="Valmiina.", bg=self.CARD, fg=self.TEXT, font=("Segoe UI", 11))
        self.status.pack(anchor="w", padx=12, pady=(12, 6))

        self.progress = ttk.Progressbar(status_card, orient="horizontal", mode="determinate", length=940)
        self.progress.pack(fill="x", padx=12, pady=(0, 12))

        pdf_card = self._card(root)
        pdf_card.pack(fill="x", padx=16, pady=(0, 12))

        self.drop_var = tk.StringVar(value="PDF: Pudota tähän (tai paina PDF → YTJ ja valitse tiedosto)")
        drop = tk.Label(pdf_card, textvariable=self.drop_var,
                        bg="#ffffff", fg=self.MUTED, font=("Segoe UI", 10),
                        padx=12, pady=10,
                        highlightthickness=1, highlightbackground=self.BORDER)
        drop.pack(fill="x", padx=12, pady=12)
        if HAS_DND:
            drop.drop_target_register(DND_FILES)
            drop.dnd_bind("<<Drop>>", self._on_drop_pdf)

        clip_card = self._card(root)
        clip_card.pack(fill="x", padx=16, pady=(0, 12))

        top = tk.Frame(clip_card, bg=self.CARD)
        top.pack(fill="x", padx=12, pady=(12, 6))

        self.strict_var = tk.BooleanVar(value=True)
        self.allow_asunto_var = tk.BooleanVar(value=True)
        self.location_boost_var = tk.BooleanVar(value=True)
        self.require_location_var = tk.BooleanVar(value=True)

        tk.Checkbutton(top, text="Tiukka parsinta (Oy/Ab/Ky/Tmi/...)", variable=self.strict_var,
                       bg=self.CARD, fg=self.TEXT, selectcolor="#fff").pack(side="left")
        tk.Checkbutton(top, text="Salli Asunto Oy", variable=self.allow_asunto_var,
                       bg=self.CARD, fg=self.TEXT, selectcolor="#fff").pack(side="left", padx=(16,0))
        tk.Checkbutton(top, text="Sijainti-boost", variable=self.location_boost_var,
                       bg=self.CARD, fg=self.TEXT, selectcolor="#fff").pack(side="left", padx=(16,0))
        tk.Checkbutton(top, text="Pakota sijainti jos löytyy", variable=self.require_location_var,
                       bg=self.CARD, fg=self.TEXT, selectcolor="#fff").pack(side="left", padx=(16,0))

        tk.Label(top, text="Max nimeä:", bg=self.CARD, fg=self.TEXT).pack(side="left", padx=(16, 6))
        self.max_names_var = tk.IntVar(value=400)
        tk.Spinbox(top, from_=10, to=5000, textvariable=self.max_names_var, width=7,
                   bg="#ffffff", fg=self.TEXT, insertbackground=self.TEXT,
                   highlightthickness=1, highlightbackground=self.BORDER).pack(side="left")

        self._btn(top, "Tyhjennä", self.clear_clipboard_text, kind="grey").pack(side="right", padx=6)

        tk.Label(clip_card, text="Liitä tähän (Ctrl+V):", bg=self.CARD, fg=self.MUTED, font=("Segoe UI", 10)).pack(anchor="w", padx=12, pady=(6, 6))

        self.clip_text = tk.Text(clip_card, height=9, wrap="word",
                                 bg="#ffffff", fg=self.TEXT, insertbackground=self.TEXT,
                                 highlightthickness=1, highlightbackground=self.BORDER)
        self.clip_text.pack(fill="x", padx=12, pady=(0, 12))

        log_card = self._card(root)
        log_card.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        tk.Label(log_card, text="Logi (vain tässä näkymässä):", bg=self.CARD, fg=self.MUTED, font=("Segoe UI", 10)).pack(anchor="w", padx=12, pady=(12, 6))

        body = tk.Frame(log_card, bg=self.CARD)
        body.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        self.listbox = tk.Listbox(body, height=16, bg="#ffffff", fg=self.TEXT,
                                  highlightthickness=1, highlightbackground=self.BORDER,
                                  selectbackground="#dbeafe")
        self.listbox.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(body, orient="vertical", command=self.listbox.yview)
        sb.pack(side="right", fill="y")
        self.listbox.configure(yscrollcommand=sb.set)

        self._ui_log("Ready.")

if __name__ == "__main__":
    App().mainloop()
